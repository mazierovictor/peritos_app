"""
TJRN - Unidades do Poder Judiciário - Web Scraper
===================================================
Extrai dados de unidades judiciais do Rio Grande do Norte a partir de:
https://tjrn.jus.br/unidades/

Filtra apenas pelos seguintes tipos de vara/juizado:
  - Vara Cível
  - Vara de Execução Fiscal e Tributária
  - Vara de Família e Sucessões
  - Juizado Especial Cível
  - Juizado Especial da Fazenda Pública
  - Vara Única

Gera uma planilha Excel com as colunas:
  Cidade | Órgão | Email

Estratégia:
  - O site é uma aplicação Next.js que expõe endpoints JSON internos
    no padrão /_next/data/{buildId}/...
  - O buildId é extraído dinamicamente da página inicial
  - Não é necessário Selenium (sem CAPTCHA / anti-bot neste site)
  - Dados salvos incrementalmente, com detecção de alterações

Dependências:
    pip install requests openpyxl beautifulsoup4
"""

import os
import re
import time
import logging
import unicodedata

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ──────────────────────────────────────────────
# Configurações
# ──────────────────────────────────────────────

BASE_URL    = "https://tjrn.jus.br"
OUTPUT_FILE = "tjrn_guia_judiciario.xlsx"
DELAY       = 1.5   # segundos entre requisições
PAGE_SIZE   = 10    # unidades por página (padrão do site)

# Mapeamento: (categoria_id, tipo_id) → nome amigável
# Extraído do site via inspeção das dropdowns
FILTER_LIST = [
    (6,  8,  "Vara Cível"),
    (6,  13, "Vara de Execução Fiscal e Tributária"),
    (6,  11, "Vara de Família e Sucessões"),
    (5,  18, "Juizado Especial Cível"),
    (5,  19, "Juizado Especial da Fazenda Pública"),
    (6,  10, "Vara Única"),
]

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/121.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json",
    "x-nextjs-data": "1",
}


# ──────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────

def normalize(text: str) -> str:
    """Remove acentos e converte para minúsculas."""
    text = unicodedata.normalize("NFD", text)
    text = text.encode("ascii", "ignore").decode("utf-8")
    return text.lower().strip()


def strip_html(html_str: str) -> str:
    """Remove tags HTML e retorna texto limpo."""
    if not html_str:
        return ""
    soup = BeautifulSoup(html_str, "html.parser")
    return " ".join(soup.get_text(separator=" ").split())


# ──────────────────────────────────────────────
# 1. Obter buildId do Next.js
# ──────────────────────────────────────────────

def get_build_id(session: requests.Session) -> str:
    """
    Carrega a página principal e extrai o buildId do script __NEXT_DATA__.
    O buildId muda a cada deploy do site.
    """
    log.info("Obtendo buildId do Next.js em %s/unidades/ …", BASE_URL)
    resp = session.get(f"{BASE_URL}/unidades/", headers={
        "User-Agent": HEADERS["User-Agent"],
        "Accept": "text/html",
    }, timeout=30)
    resp.raise_for_status()

    # Procura pelo JSON embutido na tag <script id="__NEXT_DATA__">
    soup = BeautifulSoup(resp.text, "html.parser")
    script_tag = soup.find("script", {"id": "__NEXT_DATA__"})
    if not script_tag:
        raise RuntimeError(
            "Não foi possível encontrar __NEXT_DATA__ na página. "
            "Verifique se o site está acessível."
        )

    import json
    data = json.loads(script_tag.string)
    build_id = data.get("buildId", "")
    if not build_id:
        raise RuntimeError("buildId não encontrado em __NEXT_DATA__.")

    log.info("  → buildId: %s", build_id)
    return build_id


# ──────────────────────────────────────────────
# 2. Listar unidades por filtro (paginado)
# ──────────────────────────────────────────────

def fetch_unit_list(
    session: requests.Session,
    build_id: str,
    categoria: int,
    tipo: int,
    tipo_nome: str,
) -> list[dict]:
    """
    Percorre todas as páginas do endpoint de listagem para uma combinação
    (categoria, tipoDeUnidade) e retorna lista de slugs + nome + município.

    Estrutura real da resposta (confirmada via inspeção):
      pageProps.total          → total de unidades (ex: 35)
      pageProps.currentPage    → página atual (0-based)
      pageProps.unidadesLista  → lista de hits do Elasticsearch:
        [{_index, _type, _id, _score, _source: {titulo, municipio_nome, url, emails, ...}}]
    Cada página retorna ~20 itens.
    """
    units: list[dict] = []
    page = 0
    ITEMS_PER_PAGE = 20  # o site usa 20 itens por página

    log.info("  Buscando: %s (categoria=%d, tipo=%d) …", tipo_nome, categoria, tipo)

    while True:
        url = (
            f"{BASE_URL}/_next/data/{build_id}/unidades.json"
            f"?categoria={categoria}&tipoDeUnidade={tipo}&page={page}"
        )

        for attempt in range(3):
            try:
                resp = session.get(url, headers=HEADERS, timeout=30)
                resp.raise_for_status()
                break
            except requests.RequestException as exc:
                if attempt < 2:
                    log.warning("    Tentativa %d/3 falhou: %s. Aguardando 10s…", attempt + 1, exc)
                    time.sleep(10)
                else:
                    log.error("    Falha ao buscar página %d de %s: %s", page, tipo_nome, exc)
                    return units

        try:
            data = resp.json()
        except Exception as exc:
            log.error("    Erro ao parsear JSON (pág. %d): %s", page, exc)
            break

        page_props = data.get("pageProps", {})

        # total está direto em pageProps.total (não dentro de unidadesLista)
        total_count = page_props.get("total", 0)
        total_pages = max(1, -(-total_count // ITEMS_PER_PAGE))  # ceiling division

        if page == 0:
            log.info("    → %d unidades encontradas (%d páginas).", total_count, total_pages)

        # unidadesLista é uma lista plana de hits do Elasticsearch
        unidades_lista = page_props.get("unidadesLista", [])
        if not isinstance(unidades_lista, list):
            # Fallback: se vier num formato inesperado
            unidades_lista = unidades_lista.get("data", []) if isinstance(unidades_lista, dict) else []

        if not unidades_lista:
            log.info("    → Pág. %d vazia. Fim da listagem.", page)
            break

        for item in unidades_lista:
            # Cada item é um hit do Elasticsearch: {_index, _type, _id, _score, _source: {...}}
            source = item.get("_source", item)
            title     = source.get("titulo", "") or source.get("nome", "") or ""
            slug_raw  = source.get("url", "") or source.get("slug", "") or ""
            municipio = source.get("municipio_nome", "") or source.get("municipio", "") or ""

            # Normaliza o slug: /unidades/1-vara-civel-... → 1-vara-civel-...
            if "/unidades/" in slug_raw:
                slug = slug_raw.split("/unidades/", 1)[-1].strip("/")
            else:
                slug = slug_raw.strip("/")

            if title and slug:
                units.append({
                    "titulo":    title,
                    "slug":      slug,
                    "municipio": municipio,
                    "tipo_nome": tipo_nome,
                })

        page += 1
        if page >= total_pages:
            break

        time.sleep(DELAY)

    log.info("    → Total coletado para %s: %d unidades.", tipo_nome, len(units))
    return units


# ──────────────────────────────────────────────
# 3. Obter detalhes de uma unidade (email, telefone)
# ──────────────────────────────────────────────

def fetch_unit_detail(
    session: requests.Session,
    build_id: str,
    slug: str,
) -> dict:
    """
    Busca os detalhes de uma unidade (emails, telefones, município).
    Retorna dict com chaves: nome, municipio, emails (list), telefones (list).

    Estrutura real da resposta (confirmada via inspeção):
      pageProps.unidade.nome         → nome da unidade
      pageProps.unidade.comarca      → cidade/comarca (ex: "Natal")
      pageProps.unidade.lista_emails → [{id, titulo, descricao (=email)}, ...]
      pageProps.unidade.lista_telefones → [{id, titulo, descricao, whatsapp}, ...]
    """
    url = f"{BASE_URL}/_next/data/{build_id}/unidades/{slug}.json?id={slug}"

    for attempt in range(3):
        try:
            resp = session.get(url, headers=HEADERS, timeout=30)
            resp.raise_for_status()
            break
        except requests.RequestException as exc:
            if attempt < 2:
                log.warning("    Tentativa %d/3 falhou para %s: %s. Aguardando 10s…", attempt + 1, slug, exc)
                time.sleep(10)
            else:
                log.error("    Falha ao buscar detalhes de '%s': %s", slug, exc)
                return {}

    try:
        data = resp.json()
    except Exception as exc:
        log.error("    Erro ao parsear JSON dos detalhes de '%s': %s", slug, exc)
        return {}

    unidade = data.get("pageProps", {}).get("unidade", {})
    if not unidade:
        unidade = data.get("pageProps", {})

    nome = (
        unidade.get("nome", "")
        or unidade.get("titulo", "")
        or unidade.get("name", "")
    )
    # No detalhe, a cidade/comarca está no campo "comarca" (não "municipio")
    municipio = (
        unidade.get("comarca", "")
        or unidade.get("municipio_nome", "")
        or unidade.get("municipio", "")
    )

    # Emails: lista_emails → [{id, titulo, descricao (=endereço de email)}, ...]
    raw_emails = unidade.get("lista_emails", []) or unidade.get("emails", []) or []
    emails: list[str] = []
    for e in raw_emails:
        if isinstance(e, dict):
            # Campo "descricao" contém o endereço de email
            val = e.get("descricao", "") or e.get("email", "") or e.get("value", "")
            if val and "@" in val:
                emails.append(val.strip().lower())
            elif val:
                # Pode vir com texto extra: "Atendimento: email@..."
                m = re.search(r"[\w.\-+]+@[\w.\-]+\.\w{2,}", val)
                if m:
                    emails.append(m.group(0).lower())
        elif isinstance(e, str):
            m = re.search(r"[\w.\-+]+@[\w.\-]+\.\w{2,}", e)
            if m:
                emails.append(m.group(0).lower())

    # Telefones: lista_telefones → [{id, titulo, descricao, whatsapp}, ...]
    raw_phones = unidade.get("lista_telefones", []) or unidade.get("telefones", []) or []
    phones: list[str] = []
    for p in raw_phones:
        if isinstance(p, dict):
            titulo = p.get("titulo", "")
            desc   = p.get("descricao", "") or p.get("value", "")
            if desc:
                entry = f"{titulo}: {desc}" if titulo else desc
                phones.append(entry.strip())
        elif isinstance(p, str):
            phones.append(p.strip())

    return {
        "nome":      nome,
        "municipio": municipio,
        "emails":    emails,
        "telefones": phones,
    }


# ──────────────────────────────────────────────
# 4. Planilha Excel
# ──────────────────────────────────────────────

def load_existing_data(filename: str) -> dict[str, list[dict]]:
    """Lê a planilha existente e retorna {cidade: [records]}."""
    if not os.path.exists(filename):
        return {}
    wb = openpyxl.load_workbook(filename, data_only=True)
    ws = wb.active
    data: dict[str, list[dict]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        cidade = str(row[0]).strip()
        orgao  = str(row[1] or "").strip()
        email  = str(row[2] or "").strip()
        data.setdefault(cidade, []).append({"cidade": cidade, "orgao": orgao, "email": email})
    log.info("Planilha carregada: %d cidades existentes.", len(data))
    return data


def are_records_equal(old: list[dict], new: list[dict]) -> bool:
    """Compara duas listas (ordem-independente)."""
    if len(old) != len(new):
        return False
    def key(r):
        return (r.get("orgao", ""), r.get("email", ""))
    for o, n in zip(sorted(old, key=key), sorted(new, key=key)):
        if o["orgao"] != n["orgao"] or o["email"] != n["email"]:
            return False
    return True


def init_excel_file(filename: str) -> None:
    """Cria a planilha com cabeçalho formatado."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TJRN Guia Judiciário"
    header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    header_fill  = PatternFill("solid", fgColor="003366")
    header_align = Alignment(horizontal="center", vertical="center")
    headers = ["Cidade", "Órgão", "E-mail"]
    widths  = [30, 65, 45]
    for col_idx, (hdr, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:C1"
    wb.save(filename)
    log.info("Planilha '%s' criada.", filename)


def update_city_in_excel(data: list[dict], filename: str, city_name: str) -> None:
    """Remove linhas antigas da cidade e insere os novos registros ao final."""
    if not os.path.exists(filename):
        init_excel_file(filename)
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    # Remove de baixo para cima para não alterar os índices
    for row_idx in range(ws.max_row, 1, -1):
        val = ws.cell(row=row_idx, column=1).value
        if val and str(val).strip() == city_name:
            ws.delete_rows(row_idx)
    alt_fill   = PatternFill("solid", fgColor="E8F0FE")
    plain_fill = PatternFill("solid", fgColor="FFFFFF")
    start_row  = ws.max_row + 1
    for i, rec in enumerate(data):
        r = start_row + i
        fill = alt_fill if r % 2 == 0 else plain_fill
        for col_idx, k in enumerate(["cidade", "orgao", "email"], start=1):
            cell = ws.cell(row=r, column=col_idx, value=rec[k])
            cell.fill      = fill
            cell.alignment = Alignment(vertical="center")
    ws.auto_filter.ref = f"A1:C{max(2, ws.max_row)}"
    wb.save(filename)


# ──────────────────────────────────────────────
# 5. Main
# ──────────────────────────────────────────────

def main() -> None:
    log.info("=" * 60)
    log.info("TJRN Unidades Scraper — iniciando")
    log.info("Saída: %s", os.path.abspath(OUTPUT_FILE))
    log.info("=" * 60)

    session = requests.Session()

    # Passo 1 – buildId
    try:
        build_id = get_build_id(session)
    except Exception as exc:
        log.error("Não foi possível obter o buildId: %s", exc)
        return

    # Passo 2 – planilha
    if os.path.exists(OUTPUT_FILE):
        log.info("Planilha existente encontrada — modo atualização.")
        existing_data = load_existing_data(OUTPUT_FILE)
    else:
        init_excel_file(OUTPUT_FILE)
        existing_data = {}

    # Passo 3 – iterar sobre os filtros
    total_updates  = 0
    seen_slugs     = set()  # evita duplicatas entre filtros (uma unidade pode aparecer em mais de um)

    for categoria, tipo, tipo_nome in FILTER_LIST:
        log.info("-" * 60)
        log.info("Filtro: %s", tipo_nome)
        log.info("-" * 60)

        units = fetch_unit_list(session, build_id, categoria, tipo, tipo_nome)

        if not units:
            log.info("  Nenhuma unidade encontrada para %s.", tipo_nome)
            continue

        for idx, unit in enumerate(units, start=1):
            slug     = unit["slug"]
            titulo   = unit["titulo"]
            municipio = unit["municipio"]

            log.info(
                "  [%d/%d] %s (%s)",
                idx, len(units), titulo, municipio or slug,
            )

            if slug in seen_slugs:
                log.info("    → Já processado (duplicata entre filtros). Pulando.")
                continue
            seen_slugs.add(slug)

            # Busca detalhes
            detail = fetch_unit_detail(session, build_id, slug)
            time.sleep(DELAY)

            if not detail:
                log.warning("    → Sem detalhes para '%s'.", titulo)
                continue

            # Usa nome e município do detalhe se disponíveis, ou da lista
            nome_final     = detail.get("nome") or titulo
            municipio_final = detail.get("municipio") or municipio or nome_final
            emails         = detail.get("emails", [])

            # Gera registros — se não há email, salva com email vazio
            if emails:
                records = [
                    {"cidade": municipio_final, "orgao": nome_final, "email": email}
                    for email in emails
                ]
            else:
                records = [{"cidade": municipio_final, "orgao": nome_final, "email": ""}]

            # Compara e atualiza
            old_records = existing_data.get(municipio_final, [])
            # Filtra apenas registros do mesmo órgão para comparação
            old_for_organ = [r for r in old_records if r.get("orgao") == nome_final]

            if old_for_organ and are_records_equal(old_for_organ, records):
                log.info("    → Sem alterações.")
            else:
                # Remove registros antigos do órgão específico dentro da cidade
                # e re-insere. Como temos por órgão, usamos o nome do órgão como chave.
                update_organ_in_excel(records, OUTPUT_FILE, municipio_final, nome_final)
                total_updates += 1
                if old_for_organ:
                    log.info(
                        "    → Alterado: %d → %d registro(s).",
                        len(old_for_organ), len(records),
                    )
                else:
                    log.info("    → %d registro(s) inserido(s).", len(records))

    log.info("=" * 60)
    log.info("Concluído! %d registro(s) inserido(s)/atualizado(s).", total_updates)
    log.info("Planilha: %s", os.path.abspath(OUTPUT_FILE))
    log.info("=" * 60)


def update_organ_in_excel(data: list[dict], filename: str, city_name: str, organ_name: str) -> None:
    """
    Remove linhas antigas do órgão específico na cidade e insere os novos registros.
    Mais granular que update_city_in_excel — preserva registros de outros órgãos
    da mesma cidade.
    """
    if not os.path.exists(filename):
        init_excel_file(filename)
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    # Remove de baixo para cima
    for row_idx in range(ws.max_row, 1, -1):
        c_val = ws.cell(row=row_idx, column=1).value
        o_val = ws.cell(row=row_idx, column=2).value
        if (c_val and str(c_val).strip() == city_name and
                o_val and str(o_val).strip() == organ_name):
            ws.delete_rows(row_idx)
    alt_fill   = PatternFill("solid", fgColor="E8F0FE")
    plain_fill = PatternFill("solid", fgColor="FFFFFF")
    start_row  = ws.max_row + 1
    for i, rec in enumerate(data):
        r = start_row + i
        fill = alt_fill if r % 2 == 0 else plain_fill
        for col_idx, k in enumerate(["cidade", "orgao", "email"], start=1):
            cell = ws.cell(row=r, column=col_idx, value=rec[k])
            cell.fill      = fill
            cell.alignment = Alignment(vertical="center")
    ws.auto_filter.ref = f"A1:C{max(2, ws.max_row)}"
    wb.save(filename)


if __name__ == "__main__":
    main()
