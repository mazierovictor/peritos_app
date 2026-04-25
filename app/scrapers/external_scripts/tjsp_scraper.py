"""
TJSP - E-mails Institucionais - Web Scraper
=============================================
Extrai e-mails do Tribunal de Justiça de São Paulo:
  https://www.tjsp.jus.br/CanaisComunicacao/EmailsInstitucionais

Termos buscados (mesmos do TJSC):
  - Unica
  - Cível
  - Fazenda publica
  - administração
  - Bancario

Gera uma planilha Excel com as colunas:
  Cidade | Órgão | Email

O site NÃO possui CAPTCHA — usa requests simples (sem Selenium).

Dependências:
    pip install requests openpyxl
"""

import os
import re
import time
import logging
import unicodedata

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ─────────────────────────────────────────────────
# Configurações
# ─────────────────────────────────────────────────
API_URL     = "https://www.tjsp.jus.br/CanaisComunicacao/EmailsInstitucionais/Pesquisar"
OUTPUT_FILE = "tjsp_guia_judiciario.xlsx"

# Termos a buscar (mesmos do TJSC)
SEARCH_TERMS = [
    "Unica",
    "Civel",
    "Fazenda publica",
    "administracao",
    "Bancario",
]

# Palavras-chave para filtrar órgãos relevantes (minúsculo, sem acento)
ALLOWED_ORGANS_KEYWORDS = [
    "unica",       # Vara Única
    "civel",       # Vara Cível / Cível
    "fazenda",     # Fazenda Pública
    "administrac", # Administração
    "bancario",    # Bancário
    "bancaria",
    "vara",        # qualquer vara
    "juizado",     # Juizado Especial
    "forum",
    "foro",
]

DELAY_BETWEEN_TERMS = 2   # segundos entre cada termo buscado

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# Regex para extrair e-mail
EMAIL_RE = re.compile(r'[\w.\-+]+@[\w.\-]+\.\w{2,}', re.IGNORECASE)


# ─────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────

def _normalize(text: str) -> str:
    """Remove acentos e converte para minúsculas."""
    text = unicodedata.normalize('NFD', text)
    text = text.encode('ascii', 'ignore').decode('ascii')
    return text.lower()


def is_organ_allowed(name: str) -> bool:
    """Retorna True se o nome do órgão contiver alguma palavra-chave permitida."""
    norm = _normalize(name)
    return any(kw in norm for kw in ALLOWED_ORGANS_KEYWORDS)


def _extract_email(text: str) -> str:
    """Extrai o primeiro endereço de e-mail de um texto."""
    m = EMAIL_RE.search(text or "")
    return m.group(0).lower().strip() if m else ""


# ─────────────────────────────────────────────────
# 1. Buscar na API do TJSP
# ─────────────────────────────────────────────────

def search_api(term: str, session: requests.Session) -> list[dict]:
    """
    Faz um POST na API do TJSP buscando pelo termo dado.
    Retorna a lista de resultados (dicts com os campos do JSON).
    
    A API retorna um JSON no formato:
    [
      {
        "DisplayName": "ADAMANTINA - JUIZADO ESPECIAL CIVEL",
        "Email": "adamantjec@tjsp.jus.br",
        "Nome": "ADAMANTINA - JUIZADO ESPECIAL CIVEL"
      },
      ...
    ]
    """
    headers = {
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Referer": "https://www.tjsp.jus.br/CanaisComunicacao/EmailsInstitucionais",
        "Origin": "https://www.tjsp.jus.br",
        "X-Requested-With": "XMLHttpRequest",
    }
    payload = {"nomeSetor": term}

    try:
        resp = session.post(API_URL, data=payload, headers=headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()

        # A API retorna {"data": [...]} 
        if isinstance(data, dict) and "data" in data:
            results = data["data"]
            log.info("  Termo '%s': %d resultados.", term, len(results))
            return results
        elif isinstance(data, list):
            log.info("  Termo '%s': %d resultados.", term, len(data))
            return data
        else:
            log.warning("  Termo '%s': resposta inesperada (tipo: %s).", term, type(data).__name__)
            return []

    except Exception as exc:
        log.error("  Erro ao buscar '%s': %s", term, exc)
        return []


# ─────────────────────────────────────────────────
# 2. Processar e filtrar resultados
# ─────────────────────────────────────────────────

def parse_nome_field(nome: str) -> tuple[str, str]:
    """
    O campo 'DisplayName'/'Nome' da API do TJSP tem o formato:
      "CIDADE - TIPO DA VARA"
    Ex: "ADAMANTINA - JUIZADO ESPECIAL CIVEL"
        "SAO PAULO - FORO CENTRAL CIVEL"
    
    Retorna (cidade, orgao). Se não houver ' - ', retorna ('Não informado', nome_completo).
    """
    if " - " in nome:
        partes = nome.split(" - ", 1)
        cidade = partes[0].strip().title()
        orgao  = partes[1].strip().title()
        return cidade, orgao
    else:
        return "Não informado", nome.strip().title()


def parse_results(raw_results: list[dict]) -> list[dict]:
    """
    Converte os resultados brutos da API em registros no formato
    {cidade, orgao, email} filtrando apenas órgãos relevantes.

    Campos da API TJSP:
      - DisplayName: "CIDADE - TIPO DA VARA"
      - Email:       endereço de e-mail
      - Nome:        (geralmente igual a DisplayName)
    """
    records = []
    seen = set()  # para de-duplicação

    for item in raw_results:
        if not isinstance(item, dict):
            continue

        # Nome completo: "CIDADE - TIPO DA VARA"
        nome_completo = str(item.get("DisplayName", "") or item.get("Nome", "") or "").strip()
        email = str(item.get("Email", "") or "").strip().lower()

        # Tenta extrair e-mail se não vier no campo Email
        if not email:
            for field_val in item.values():
                cand = _extract_email(str(field_val))
                if cand:
                    email = cand
                    break

        # Valida e-mail
        if not email or "@" not in email:
            continue

        # Extrai cidade e órgão do campo DisplayName
        cidade, orgao = parse_nome_field(nome_completo)

        # Filtra por órgão relevante
        if not is_organ_allowed(nome_completo):
            log.debug("  [filtrado] %s", nome_completo)
            continue

        # De-duplicação
        key = (cidade.lower(), orgao.lower(), email)
        if key in seen:
            continue
        seen.add(key)

        records.append({
            "cidade": cidade,
            "orgao":  orgao,
            "email":  email,
        })

    return records


# ─────────────────────────────────────────────────
# 3. Gerenciar planilha Excel
# ─────────────────────────────────────────────────

def load_existing_data(filename: str) -> dict[str, list[dict]]:
    """Lê a planilha e devolve {cidade: [records]}."""
    if not os.path.exists(filename):
        return {}

    wb = openpyxl.load_workbook(filename, data_only=True)
    ws = wb.active
    data: dict[str, list[dict]] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        cidade = str(row[0]).strip()
        orgao  = str(row[1] or "").strip()
        email  = str(row[2] or "").strip()
        data.setdefault(cidade, []).append(
            {"cidade": cidade, "orgao": orgao, "email": email}
        )

    log.info("Planilha carregada: %d cidades com dados existentes.", len(data))
    return data


def are_records_equal(old: list[dict], new: list[dict]) -> bool:
    """Compara duas listas de registros (ordem-independente)."""
    if len(old) != len(new):
        return False

    def key(r):
        return (r.get("orgao", ""), r.get("email", ""))

    for o, n in zip(sorted(old, key=key), sorted(new, key=key)):
        if o["orgao"] != n["orgao"] or o["email"] != n["email"]:
            return False
    return True


def init_excel_file(filename: str) -> None:
    """Cria planilha com cabeçalho formatado."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TJSP Emails"

    header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    header_fill  = PatternFill("solid", fgColor="1B3A5C")  # Azul escuro TJSP
    header_align = Alignment(horizontal="center", vertical="center")
    headers = ["Cidade", "Órgão", "E-mail"]
    widths  = [40, 55, 45]

    for col_idx, (hdr, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:C1"
    wb.save(filename)
    log.info("Planilha '%s' criada.", filename)


def upsert_records_in_excel(
    all_records: list[dict],
    filename: str,
    existing_data: dict[str, list[dict]],
) -> int:
    """
    Insere ou atualiza registros na planilha.
    Agrupa por cidade, compara com dados existentes, e salva incrementalmente.
    Retorna o número de cidades atualizadas.
    """
    if not os.path.exists(filename):
        init_excel_file(filename)

    # Agrupa novos registros por cidade
    by_city: dict[str, list[dict]] = {}
    for rec in all_records:
        by_city.setdefault(rec["cidade"], []).append(rec)

    updates = 0
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    alt_fill   = PatternFill("solid", fgColor="E8F0F8")
    plain_fill = PatternFill("solid", fgColor="FFFFFF")

    for cidade_label, new_recs in sorted(by_city.items()):
        old_recs = existing_data.get(cidade_label, [])

        if old_recs and are_records_equal(old_recs, new_recs):
            log.info("  [%s] Sem alterações (%d registros).", cidade_label, len(new_recs))
            continue

        # Remove linhas antigas desta cidade
        for row_idx in range(ws.max_row, 1, -1):
            val = ws.cell(row=row_idx, column=1).value
            if val and str(val).strip() == cidade_label:
                ws.delete_rows(row_idx)

        # Insere novos registros
        start_row = ws.max_row + 1
        for i, rec in enumerate(new_recs):
            r = start_row + i
            fill = alt_fill if r % 2 == 0 else plain_fill
            for col_idx, fld in enumerate(["cidade", "orgao", "email"], start=1):
                cell = ws.cell(row=r, column=col_idx, value=rec[fld])
                cell.fill      = fill
                cell.alignment = Alignment(vertical="center")

        if old_recs:
            log.info("  [%s] Atualizado: %d → %d registros.", cidade_label, len(old_recs), len(new_recs))
        else:
            log.info("  [%s] Inserido: %d registros.", cidade_label, len(new_recs))
        updates += 1

    ws.auto_filter.ref = f"A1:C{max(2, ws.max_row)}"
    wb.save(filename)
    return updates


# ─────────────────────────────────────────────────
# 4. Main
# ─────────────────────────────────────────────────

def main():
    log.info("=" * 65)
    log.info("TJSP - E-mails Institucionais — Scraper")
    log.info("=" * 65)
    log.info("Termos a buscar: %s", ", ".join(SEARCH_TERMS))
    log.info("Saída: %s", os.path.abspath(OUTPUT_FILE))
    log.info("=" * 65)

    # Planilha existente ou nova
    if os.path.exists(OUTPUT_FILE):
        log.info("Planilha existente encontrada — modo atualização.")
        existing_data = load_existing_data(OUTPUT_FILE)
    else:
        init_excel_file(OUTPUT_FILE)
        existing_data = {}

    # Sessão HTTP (sem necessidade de Selenium — não tem CAPTCHA)
    session = requests.Session()
    session.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/121.0.0.0 Safari/537.36"
        )
    })

    all_records: list[dict] = []
    seen_global = set()  # de-duplicação global entre termos

    for term in SEARCH_TERMS:
        log.info("Buscando: '%s'", term)
        raw_results = search_api(term, session)

        # Processa resultados
        parsed = parse_results(raw_results)

        # De-duplicação global (entre termos diferentes)
        new_parsed = []
        for rec in parsed:
            key = (rec["cidade"].lower(), rec["orgao"].lower(), rec["email"])
            if key not in seen_global:
                seen_global.add(key)
                new_parsed.append(rec)

        log.info("  → %d registros válidos após filtragem (%d novos).",
                 len(parsed), len(new_parsed))
        all_records.extend(new_parsed)

        if term != SEARCH_TERMS[-1]:
            time.sleep(DELAY_BETWEEN_TERMS)

    # Salvar na planilha
    if all_records:
        log.info("Total de registros coletados: %d", len(all_records))
        updates = upsert_records_in_excel(all_records, OUTPUT_FILE, existing_data)
        log.info("=" * 65)
        log.info("Concluído! %d cidade(s) inseridas/atualizadas.", updates)
    else:
        log.warning("Nenhum registro coletado. Verifique a conectividade e tente novamente.")

    log.info("Planilha: %s", os.path.abspath(OUTPUT_FILE))
    log.info("=" * 65)


if __name__ == "__main__":
    main()
