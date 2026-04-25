"""
TJMT - Canais Permanentes de Acesso - Web Scraper
===================================================
Extrai dados de contato do portal:
  https://canaispermanentesdeacesso.tjmt.jus.br/pagina/8

Seções alvo:
  1. Núcleo de Justiça 4.0 do Juiz de Garantias
  2. Núcleo de Justiça 4.0 - Execução Fiscal Estadual
  3. Núcleo de Justiça Digital dos Juizados Especiais
  + Todas as Comarcas

Dentro de cada Comarca, captura apenas:
  - Diretoria do Fórum
  - Distribuição e Protocolo (Central de Distribuição)
  - Varas Cíveis (Secretaria + Gabinete)
  - Vara Única (Secretaria + Gabinete)

Gera planilha Excel: tjmt_guia_judiciario.xlsx
  Colunas: Cidade | Órgão | E-mail

Estratégia:
  - Selenium carrega o Angular SPA
  - Aguarda renderização completa
  - page_source é passado para BeautifulSoup para parse estático
  - Angular renderiza todo o HTML no DOM (hidden com CSS), sem necessidade de clicar

Dependências:
    pip install selenium openpyxl beautifulsoup4
    ChromeDriver compatível com o Chrome instalado
"""

import os
import re
import time
import logging
import unicodedata

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ──────────────────────────────────────────────
# Configurações
# ──────────────────────────────────────────────

URL         = "https://canaispermanentesdeacesso.tjmt.jus.br/pagina/8"
OUTPUT_FILE = "tjmt_guia_judiciario.xlsx"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ──────────────────────────────────────────────
# Filtros
# ──────────────────────────────────────────────

# Palavras-chave que identificam itens RAIZ que devem ser tratados como "Núcleos"
# (captura todos os emails de todos os sub-itens, sem filtro adicional)
NUCLEO_KEYWORDS = [
    "nucleo de justica 4.0 do juiz de garantias",
    "nucleo de justica 4.0 - execucao fiscal estadual",
    "execucao fiscal estadual",
    "nucleo de justica digital dos juizados especiais",
    "vara especializada da fazenda publica",
    "vara especializada de familia e sucessoes",
    "vara especializada em direito bancario",
    "juizado especial civel",
    "juizado especial da fazenda publica",
    "vara especializada de execucao fiscal",
]

# ─── Override pela UI (não altera lógica; só substitui a lista se houver config salva) ───
try:
    import json as _json_ui
    with open("scraper_config.json", encoding="utf-8") as _f_ui:
        _UI_CFG = _json_ui.load(_f_ui)
    if isinstance(_UI_CFG.get("palavras_chave"), list) and _UI_CFG["palavras_chave"]:
        NUCLEO_KEYWORDS = [str(x) for x in _UI_CFG["palavras_chave"]]
except Exception:
    pass

# Sub-seções desejadas dentro das Comarcas
COMARCA_INCLUDE = [
    "diretoria do forum",
    "diretoria do foro",
    "distribuicao",
    "protocolo",
    "vara civel",
    "varas civeis",
    "vara unica",
    "vara especializada da fazenda publica",
    "vara especializada de familia e sucessoes",
    "vara especializada em direito bancario",
    "vara especializada de execucao fiscal",
    "juizado especial civel",
    "juizado especial da fazenda publica",
    "secretaria",
    "gabinete",
]

# Sub-seções a ignorar (varas não-cíveis)
COMARCA_EXCLUDE = [
    "criminal",
    "criminais",
    "execucoes penais",
    "execucao penal",
    "juizado especial criminal",
    "infancia",
    "família",
    "familia",
    "registros publicos",
    "orfaos",
    "tutela",
    "violencia domestica",
]


# ──────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────

def norm(text: str) -> str:
    """Normaliza texto: minúsculas, sem acentos."""
    t = unicodedata.normalize("NFD", text)
    t = t.encode("ascii", "ignore").decode("utf-8")
    return t.lower().strip()


def extract_emails(text: str) -> list:
    return re.findall(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", text)


def is_nucleo(title: str) -> bool:
    """Verifica se o título bate com algum item raiz da lista NUCLEO_KEYWORDS."""
    n = norm(title)
    for kw in NUCLEO_KEYWORDS:
        if kw in n:
            return True
    return False


def include_comarca_section(title: str) -> bool:
    """Retorna True se a sub-seção da Comarca deve ser incluída."""
    n = norm(title)
    for ex in COMARCA_EXCLUDE:
        if ex in n:
            return False
    for inc in COMARCA_INCLUDE:
        if inc in n:
            return True
    return False


def is_vara_relevante(title: str) -> bool:
    """Retorna True se é uma Vara/Juizado relevante (inclui tudo abaixo via force_include)."""
    n = norm(title)
    return (
        "vara civel" in n
        or "varas civeis" in n
        or "vara unica" in n
        or "vara especializada" in n
        or "juizado especial civel" in n
        or "juizado especial da fazenda" in n
    )


# ──────────────────────────────────────────────
# Planilha Excel
# ──────────────────────────────────────────────

def init_excel(filename: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TJMT Guia Judiciário"
    hfont  = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    hfill  = PatternFill("solid", fgColor="003366")
    halign = Alignment(horizontal="center", vertical="center")
    headers = ["Cidade", "Órgão", "E-mail"]
    widths  = [30, 65, 45]
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = halign
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:C1"
    wb.save(filename)
    log.info("Planilha '%s' criada.", filename)


def append_records(records: list, filename: str) -> None:
    """Adiciona registros ao final da planilha, evitando duplicatas exatas."""
    if not os.path.exists(filename):
        init_excel(filename)

    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[1] and row[2]:
            existing.add((str(row[1]).strip().lower(), str(row[2]).strip().lower()))

    alt_fill   = PatternFill("solid", fgColor="E8F0FE")
    plain_fill = PatternFill("solid", fgColor="FFFFFF")

    added = 0
    for rec in records:
        key = (rec["orgao"].strip().lower(), rec["email"].strip().lower())
        if key in existing:
            continue
        existing.add(key)
        r = ws.max_row + 1
        fill = alt_fill if r % 2 == 0 else plain_fill
        for col_idx, k in enumerate(["cidade", "orgao", "email"], start=1):
            cell = ws.cell(row=r, column=col_idx, value=rec[k])
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")
        added += 1

    ws.auto_filter.ref = f"A1:C{max(2, ws.max_row)}"
    wb.save(filename)
    log.info("  → %d registro(s) adicionado(s) à planilha.", added)


# ──────────────────────────────────────────────
# Selenium
# ──────────────────────────────────────────────

def setup_driver() -> webdriver.Chrome:
    opts = Options()
    # Em container/servidor sempre força headless (CHROME_BIN é setada no Dockerfile)
    if os.environ.get("CHROME_BIN"):
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1400,900")
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
    )

    chrome_bin = os.environ.get("CHROME_BIN")
    chromedriver_path = os.environ.get("CHROMEDRIVER_PATH")
    if chrome_bin:
        opts.binary_location = chrome_bin

    proxy = os.environ.get("HTTPS_PROXY") or os.environ.get("HTTP_PROXY")
    if proxy:
        opts.add_argument(f"--proxy-server={proxy}")

    if chromedriver_path:
        from selenium.webdriver.chrome.service import Service as _ChromeService
        driver = webdriver.Chrome(service=_ChromeService(chromedriver_path), options=opts)
    else:
        driver = webdriver.Chrome(options=opts)
    return driver


def get_page_source(driver: webdriver.Chrome) -> str:
    """Carrega a página e retorna o page_source após o Angular renderizar."""
    log.info("Abrindo página: %s", URL)
    driver.get(URL)

    # Aguarda o Angular renderizar os primeiros itens da árvore
    WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "app-tree-view-item"))
    )
    log.info("Componentes app-tree-view-item detectados. Aguardando renderização completa...")

    # Scroll para baixo progressivamente para forçar a renderização de todos os componentes
    # (Angular pode usar virtual scrolling)
    last_height = driver.execute_script("return document.body.scrollHeight")
    for _ in range(20):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1.5)
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height

    # Volta ao topo
    driver.execute_script("window.scrollTo(0, 0);")
    time.sleep(2)

    html = driver.page_source
    log.info("page_source obtido: %d bytes.", len(html))
    return html


# ──────────────────────────────────────────────
# Parse BeautifulSoup da árvore Angular
# ──────────────────────────────────────────────

def get_item_title(item_tag) -> str:
    """Retorna o título do app-tree-view-item (texto do h4)."""
    h4 = item_tag.find("h4", recursive=False)
    if not h4:
        # Tenta dentro de uma div direta
        for child_div in item_tag.find_all("div", recursive=False):
            h4 = child_div.find("h4", recursive=False)
            if h4:
                break
    if not h4:
        h4 = item_tag.find("h4")
    return h4.get_text(strip=True) if h4 else ""


def get_direct_children_items(item_tag) -> list:
    """
    Retorna os app-tree-view-item filhos diretos de item_tag.
    Evita capturar children de níveis mais profundos.
    """
    direct = []
    # O conteúdo do accordion normalmente está em .conteudo ou .accordion-item-content
    content_div = item_tag.find("div", class_=lambda c: c and (
        "conteudo" in c or "accordion-item-content" in c or "content" in c
    ), recursive=False)

    if content_div is None:
        # Tenta procurar dentro de divs filhas diretas
        for div in item_tag.find_all("div", recursive=False):
            content_div = div.find("div", class_=lambda c: c and (
                "conteudo" in c or "accordion-item-content" in c
            ), recursive=False)
            if content_div:
                break

    if content_div:
        for child in content_div.find_all("app-tree-view-item", recursive=False):
            direct.append(child)
    else:
        # Fallback: qualquer filho direto
        for child in item_tag.find_all("app-tree-view-item", recursive=False):
            direct.append(child)

    return direct


def get_direct_text_emails(item_tag) -> list:
    """
    Extrai emails do conteúdo DIRETO do item (não de sub-itens).
    Copia a tag, remove filhos app-tree-view-item e extrai o texto.
    """
    # Encontra o content_div
    content_div = item_tag.find("div", class_=lambda c: c and (
        "conteudo" in c or "accordion-item-content" in c or "content" in c
    ))

    if not content_div:
        return []

    # Copia o content_div e remove sub-itens
    content_copy = BeautifulSoup(str(content_div), "html.parser")
    for sub in content_copy.find_all("app-tree-view-item"):
        sub.decompose()

    text = content_copy.get_text(separator=" ")
    emails = extract_emails(text)

    # Também verifica links mailto:
    for a in content_copy.find_all("a", href=lambda h: h and "mailto:" in h):
        e = a["href"].replace("mailto:", "").strip()
        if e:
            emails.append(e)

    return list(set(emails))


def get_all_emails_in_subtree(item_tag) -> list:
    """
    Extrai TODOS os emails de toda a subárvore do item (incluindo sub-itens
    em qualquer nível de aninhamento: Secretaria, Gabinete, etc.).
    Usado para Varas relevantes, onde queremos capturar tudo abaixo.
    """
    emails = []

    # Emails via links mailto: em qualquer profundidade
    for a in item_tag.find_all("a", href=lambda h: h and "mailto:" in h):
        e = a["href"].replace("mailto:", "").strip()
        if e and "@" in e:
            emails.append(e)

    # Emails via texto (regex) em toda a árvore
    text = item_tag.get_text(separator=" ")
    emails.extend(extract_emails(text))

    return list(set(emails))


# ──────────────────────────────────────────────
# Processamento da Árvore
# ──────────────────────────────────────────────

def process_nucleo_item(item_tag) -> list:
    """
    Processa recursivamente um nó de Núcleo.
    Captura todos os emails em todos os sub-itens.
    """
    records = []
    nucleo_title = get_item_title(item_tag)

    def walk(node, path):
        title = get_item_title(node)
        current_path = f"{path} > {title}".strip(" > ") if title else path

        emails = get_direct_text_emails(node)
        for email in emails:
            records.append({
                "cidade": nucleo_title,
                "orgao": current_path or nucleo_title,
                "email": email,
            })
            log.info("    [Núcleo] %s → %s", current_path, email)

        for child in get_direct_children_items(node):
            walk(child, current_path)

    # Emails diretos do Núcleo raiz
    for email in get_direct_text_emails(item_tag):
        records.append({
            "cidade": nucleo_title,
            "orgao": nucleo_title,
            "email": email,
        })

    # Filhos do Núcleo
    for child in get_direct_children_items(item_tag):
        walk(child, "")

    return records


def process_comarca_item(item_tag) -> list:
    """
    Processa um nó de Comarca.
    Para seções de endereço (Diretoria, Distribuição): extrai emails diretos.
    Para varas relevantes (Cível, Única, Especializadas, Juizados):
      coleta TODOS os emails da subárvore inteira (Secretaria + Gabinete + qualquer nível)
      e grava sob o path da própria Vara — sem recursão adicional nos filhos.
    """
    records = []
    comarca_title = get_item_title(item_tag)

    def walk(node, path, force_include=False):
        title = get_item_title(node)
        current_path = f"{path} > {title}".strip(" > ") if title else path
        n_title = norm(title)

        # ── Dentro de vara relevante (force_include=True) ────────────────────
        # Captura todos os emails da subárvore e grava sob o path da vara pai.
        # Não recursamos mais: a vara pai já coletou tudo.
        if force_include:
            emails = get_all_emails_in_subtree(node)
            for email in emails:
                records.append({
                    "cidade": comarca_title,
                    "orgao": path,          # ← path da VARA, não do sub-item
                    "email": email,
                })
                log.info("    [Comarca] %s > %s → %s", comarca_title, path, email)
            # Não recursamos: já coletamos tudo da subárvore
            return

        # ── Seção de nível 1 (filhos diretos da Comarca) ─────────────────────
        # Checa exclusão primeiro
        is_excl = any(ex in n_title for ex in COMARCA_EXCLUDE)
        if is_excl:
            return  # ignora esta seção e todos os filhos

        vara_rel = is_vara_relevante(title)

        if vara_rel:
            # É uma Vara relevante: captura TODA a subárvore (sec + gab + ...)
            # sob o path desta vara e não recursamos nos filhos.
            emails = get_all_emails_in_subtree(node)
            for email in emails:
                records.append({
                    "cidade": comarca_title,
                    "orgao": current_path,
                    "email": email,
                })
                log.info("    [Comarca] %s > %s → %s", comarca_title, current_path, email)
            # Não recursamos: a subárvore inteira já foi consumida
            return

        # Seção de apoio (Diretoria, Distribuição, Protocolo, etc.)
        include = include_comarca_section(title)
        if include:
            emails = get_direct_text_emails(node)
            for email in emails:
                records.append({
                    "cidade": comarca_title,
                    "orgao": current_path,
                    "email": email,
                })
                log.info("    [Comarca] %s > %s → %s", comarca_title, current_path, email)

        # Continua percorrendo filhos sem force_include
        for child in get_direct_children_items(node):
            walk(child, current_path, force_include=False)

    # Processa filhos da comarca (não a comarca em si)
    for child in get_direct_children_items(item_tag):
        walk(child, "", force_include=False)

    return records


def process_tree(html: str) -> list:
    """Parseia o HTML e processa todos os itens raiz da árvore."""
    soup = BeautifulSoup(html, "html.parser")

    # Encontra todos os app-tree-view-item raiz (sem parent app-tree-view-item)
    all_items = soup.find_all("app-tree-view-item")
    root_items = [
        item for item in all_items
        if item.parent is not None and item.parent.name != "app-tree-view-item"
        and not item.find_parent("app-tree-view-item")
    ]

    log.info("Total de itens raiz encontrados: %d", len(root_items))

    all_records = []

    for item in root_items:
        title = get_item_title(item)
        if not title:
            continue

        log.info("Processando: %s", title)

        if is_nucleo(title):
            recs = process_nucleo_item(item)
            log.info("  → %d registro(s) do Núcleo.", len(recs))
            all_records.extend(recs)
        else:
            # Trata como Comarca
            recs = process_comarca_item(item)
            if recs:
                log.info("  → %d registro(s) da Comarca.", len(recs))
            all_records.extend(recs)

    return all_records


# ──────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────

def main() -> None:
    log.info("=" * 60)
    log.info("TJMT Canais Permanentes de Acesso - Scraper")
    log.info("Saida: %s", os.path.abspath(OUTPUT_FILE))
    log.info("=" * 60)

    driver = setup_driver()
    try:
        html = get_page_source(driver)
    finally:
        log.info("Fechando navegador...")
        driver.quit()

    if not html:
        log.error("Nenhum HTML obtido. Encerrando.")
        return

    log.info("-" * 60)
    log.info("Processando HTML com BeautifulSoup...")

    records = process_tree(html)

    log.info("-" * 60)
    log.info("Total de registros extraidos: %d", len(records))

    if not records:
        log.warning("Nenhum registro encontrado. Verifique a estrutura HTML e os filtros.")
        return

    append_records(records, OUTPUT_FILE)

    log.info("=" * 60)
    log.info("Concluido! Planilha: %s", os.path.abspath(OUTPUT_FILE))
    log.info("=" * 60)


if __name__ == "__main__":
    main()
