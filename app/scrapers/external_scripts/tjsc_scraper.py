"""
TJSC - Consulta de E-mails - Web Scraper
==========================================
Extrai e-mails do sistema CEM (Consulta de E-mails) do TJSC:
  https://app.tjsc.jus.br/tjsc-consulta-email/#/

Termos buscados:
  - Unica
  - Cível
  - Fazenda publica
  - administração
  - Bancario

Gera uma planilha Excel com as colunas:
  Cidade | Órgão | Email

Estratégia de CAPTCHA:
  O site usa reCAPTCHA Enterprise. O script abre o browser de forma VISÍVEL,
  aguarda o usuário resolver o CAPTCHA manualmente (apenas uma vez),
  captura o token gerado e o usa para todas as requisições à API REST.

Dependências:
    pip install selenium webdriver-manager openpyxl requests
"""

import os
import re
import time
import logging
import unicodedata

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

try:
    from webdriver_manager.chrome import ChromeDriverManager
    USE_WDM = True
except ImportError:
    USE_WDM = False

# ─────────────────────────────────────────────────
# Configurações
# ─────────────────────────────────────────────────
BASE_URL    = "https://app.tjsc.jus.br/tjsc-consulta-email/#/"
API_URL     = "https://app.tjsc.jus.br/tjsc-consulta-email/rest/publico/email"
OUTPUT_FILE = "tjsc_guia_judiciario.xlsx"

# Termos a buscar
SEARCH_TERMS = [
    "Unica",
    "Civel",
    "Fazenda publica",
    "administracao",
    "Bancario",
]

# ─── Override pela UI (não altera lógica; só substitui a lista se houver config salva) ───
try:
    import json as _json_ui
    with open("scraper_config.json", encoding="utf-8") as _f_ui:
        _UI_CFG = _json_ui.load(_f_ui)
    if isinstance(_UI_CFG.get("palavras_chave"), list) and _UI_CFG["palavras_chave"]:
        SEARCH_TERMS = [str(x) for x in _UI_CFG["palavras_chave"]]
except Exception:
    pass

# Palavras-chave para filtrar órgãos relevantes (minúsculo, sem acento)
# Correspondentes aos 5 termos de busca no TJSC
ALLOWED_ORGANS_KEYWORDS = [
    "unica",       # Vara Única
    "civel",       # Vara Cível / Cível
    "fazenda",     # Fazenda Pública
    "administrac", # Administração
    "bancario",    # Bancário
    "bancaria",
    "vara",        # qualquer vara
    "juizado",     # Juizado Especial
    "forum",
    "foro",
]

DELAY_BETWEEN_TERMS = 2   # segundos entre cada termo buscado
CAPTCHA_WAIT_TIMEOUT = 300  # segundos para o usuário resolver o CAPTCHA (5 min)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# Regex para extrair e-mail
EMAIL_RE = re.compile(r'[\w.\-+]+@[\w.\-]+\.\w{2,}', re.IGNORECASE)


# ─────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────

def _normalize(text: str) -> str:
    """Remove acentos e converte para minúsculas."""
    text = unicodedata.normalize('NFD', text)
    text = text.encode('ascii', 'ignore').decode('ascii')
    return text.lower()


def is_organ_allowed(name: str) -> bool:
    """Retorna True se o nome do órgão contiver alguma palavra-chave permitida."""
    norm = _normalize(name)
    return any(kw in norm for kw in ALLOWED_ORGANS_KEYWORDS)


def _extract_email(text: str) -> str:
    """Extrai o primeiro endereço de e-mail de um texto."""
    m = EMAIL_RE.search(text or "")
    return m.group(0).lower().strip() if m else ""


# ─────────────────────────────────────────────────
# 1. Inicializar o WebDriver (visível para CAPTCHA)
# ─────────────────────────────────────────────────

def create_driver() -> webdriver.Chrome:
    opts = Options()
    # NÃO headless — precisa ser visível para o usuário resolver o CAPTCHA
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1200,800")
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/121.0.0.0 Safari/537.36"
    )

    if USE_WDM:
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=opts)
    else:
        driver = webdriver.Chrome(options=opts)

    driver.implicitly_wait(5)
    return driver


# ─────────────────────────────────────────────────
# 2. Obter token reCAPTCHA via interação humana
# ─────────────────────────────────────────────────

def get_recaptcha_token(driver: webdriver.Chrome) -> str:
    """
    Abre a página do TJSC, aguarda o usuário resolver o reCAPTCHA manualmente,
    e captura o token gerado pelo reCAPTCHA.
    
    Retorna o token (string) ou '' se não conseguir capturar.
    """
    log.info("Abrindo o site do TJSC...")
    driver.get(BASE_URL)
    time.sleep(3)

    print("\n" + "="*65)
    print("  ATENÇÃO — INTERVENÇÃO HUMANA NECESSÁRIA")
    print("="*65)
    print("  O site do TJSC usa reCAPTCHA. Siga os passos:")
    print("  1. Olhe para a janela do Chrome que foi aberta")
    print("  2. Clique em 'Não sou um robô' e resolva o desafio")
    print("  3. Após resolver, o script continuará automaticamente")
    print(f"  (Você tem {CAPTCHA_WAIT_TIMEOUT // 60} minutos para resolver)")
    print("="*65 + "\n")

    # Aguarda o token do reCAPTCHA aparecer na página
    # O reCAPTCHA Enterprise armazena o token em window.grecaptchaToken
    # ou em textarea[name=g-recaptcha-response]
    token = ""
    deadline = time.time() + CAPTCHA_WAIT_TIMEOUT

    while time.time() < deadline:
        try:
            # Tenta capturar o token via textarea (padrão reCAPTCHA v2)
            token = driver.execute_script(
                """
                try {
                    var ta = document.querySelector('textarea[name="g-recaptcha-response"]');
                    if (ta && ta.value && ta.value.length > 20) return ta.value;
                    
                    // Tenta reCAPTCHA Enterprise
                    var frames = document.querySelectorAll('iframe[src*="recaptcha"]');
                    if (window.___grecaptcha_cfg) {
                        var clients = window.___grecaptcha_cfg.clients;
                        if (clients) {
                            for (var k in clients) {
                                var c = clients[k];
                                for (var p in c) {
                                    if (c[p] && c[p].response && c[p].response.length > 20) {
                                        return c[p].response;
                                    }
                                }
                            }
                        }
                    }
                    return null;
                } catch(e) { return null; }
                """
            )
            if token and len(token) > 20:
                log.info("Token reCAPTCHA capturado com sucesso! (%.0f chars)", len(token))
                break

        except Exception as e:
            log.debug("Aguardando token... %s", e)

        time.sleep(2)

    if not token:
        log.warning(
            "Não foi possível capturar o token automaticamente. "
            "Tentando estratégia alternativa: interceptando a requisição..."
        )
        # Estratégia alternativa: preenche o campo Nome com o primeiro termo
        # e intercepta a requisição via log de performance
        try:
            # Habilita log de performance para capturar requests
            caps = driver.desired_capabilities
            log.warning("Token não capturado. O script tentará a busca sem ele.")
        except Exception:
            pass

    return token or ""


# ─────────────────────────────────────────────────
# 3. Fazer buscas na API com o token
# ─────────────────────────────────────────────────

def search_api(term: str, recaptcha_token: str, session: requests.Session) -> list[dict]:
    """
    Faz um POST na API do TJSC buscando pelo termo dado.
    Retorna a lista de resultados (dicts com os campos do JSON).
    """
    headers = {
        "Content-Type": "application/json;charset=UTF-8",
        "Accept": "application/json, text/plain, */*",
        "Referer": "https://app.tjsc.jus.br/tjsc-consulta-email/",
        "Origin": "https://app.tjsc.jus.br",
        "recaptchaResponse": recaptcha_token,
    }
    payload = {"nome": term}

    try:
        resp = session.post(API_URL, json=payload, headers=headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        usuarios = data.get("dadosUsuarios", [])
        log.info("  Termo '%s': %d resultados (código=%s)",
                 term, len(usuarios), data.get("codigoResultado", "?"))
        return usuarios
    except Exception as exc:
        log.error("  Erro ao buscar '%s': %s", term, exc)
        return []


# ─────────────────────────────────────────────────
# 4. Busca via Selenium (fallback se API falhar)
# ─────────────────────────────────────────────────

def search_via_browser(driver: webdriver.Chrome, term: str) -> list[dict]:
    """
    Faz a busca diretamente no browser (fallback caso o token da API expire).
    Preenche o campo Nome e intercepta a resposta via JavaScript.
    """
    log.info("  Buscando '%s' via browser...", term)
    results = []

    try:
        # Garante que estamos na página
        if BASE_URL not in driver.current_url:
            driver.get(BASE_URL)
            time.sleep(3)

        # Intercepta as chamadas XHR para capturar a resposta
        driver.execute_script("""
            window._tjsc_last_response = null;
            var originalOpen = XMLHttpRequest.prototype.open;
            var originalSend = XMLHttpRequest.prototype.send;
            XMLHttpRequest.prototype.open = function(method, url) {
                this._url = url;
                return originalOpen.apply(this, arguments);
            };
            XMLHttpRequest.prototype.send = function(body) {
                var xhr = this;
                xhr.addEventListener('load', function() {
                    if (xhr._url && xhr._url.indexOf('email') !== -1 && xhr.status === 200) {
                        try {
                            window._tjsc_last_response = JSON.parse(xhr.responseText);
                        } catch(e) {}
                    }
                });
                return originalSend.apply(this, arguments);
            };
        """)

        # Limpa e preenche o campo Nome
        wait = WebDriverWait(driver, 15)
        nome_input = wait.until(EC.presence_of_element_located(
            (By.CSS_SELECTOR, "input[ng-model*='nome'], input[placeholder*='nome'], input[placeholder*='Nome'], #nome")
        ))
        nome_input.clear()
        nome_input.send_keys(term)
        time.sleep(1)

        # Clica em Consultar
        consultar_btn = driver.find_element(
            By.CSS_SELECTOR, "button[ng-click*='consultar'], button[ng-click*='pesquisar'], .btn-primary"
        )
        driver.execute_script("arguments[0].click();", consultar_btn)
        time.sleep(4)

        # Captura a resposta interceptada
        response_data = driver.execute_script("return window._tjsc_last_response;")
        if response_data and "dadosUsuarios" in response_data:
            results = response_data["dadosUsuarios"]
            log.info("    → %d resultados via browser para '%s'", len(results), term)

    except Exception as exc:
        log.error("  Erro no fallback de browser para '%s': %s", term, exc)

    return results


# ─────────────────────────────────────────────────
# 5. Processar e filtrar resultados
# ─────────────────────────────────────────────────

def parse_nome_field(nome: str) -> tuple[str, str]:
    """
    O campo 'nome' da API do TJSC tem o formato:
      "Cidade - Tipo da Vara"
    Ex: "Abelardo Luz - Vara Única"
        "Itaiópolis - Vara Única"
        "Florianópolis - 1ª Vara Cível"
    
    Retorna (cidade, orgao). Se não houver ' - ', retorna ('Não informado', nome_completo).
    """
    if " - " in nome:
        partes = nome.split(" - ", 1)
        cidade = partes[0].strip().title()
        orgao  = partes[1].strip()
        return cidade, orgao
    else:
        return "Não informado", nome


def parse_results(raw_results: list[dict]) -> list[dict]:
    """
    Converte os resultados brutos da API em registros no formato
    {cidade, orgao, email} filtrando apenas órgãos relevantes.

    Campos da API TJSC:
      - nome:    "Cidade - Tipo da Vara"  (ex: "Abelardo Luz - Vara Única")
      - email:   endereço de e-mail
      - lotacao: "CIDADE - VARA ÚNICA" (redundante, usado como fallback)
      - matricula: null (irrelevante)
    """
    records = []
    seen = set()  # para de-duplicação

    for item in raw_results:
        if not isinstance(item, dict):
            continue

        # Nome completo: "Cidade - Tipo da Vara"
        nome_completo = str(item.get("nome", "") or "").strip()
        email = str(item.get("email", "") or "").strip().lower()

        # Tenta extrair e-mail se não vier no campo email
        if not email:
            for field_val in item.values():
                cand = _extract_email(str(field_val))
                if cand:
                    email = cand
                    break

        # Valida e-mail
        if not email or "@" not in email:
            continue

        # Extrai cidade e órgão do campo 'nome'
        cidade, orgao = parse_nome_field(nome_completo)

        # Filtra por órgão relevante (verifica contra o nome completo para
        # capturar casos onde o termo está na parte cidade, ex: lotações administrativas)
        if not is_organ_allowed(nome_completo):
            log.debug("  [filtrado] %s", nome_completo)
            continue

        # De-duplicação
        key = (cidade.lower(), orgao.lower(), email)
        if key in seen:
            continue
        seen.add(key)

        records.append({
            "cidade": cidade,
            "orgao":  orgao,
            "email":  email,
        })

    return records


# ─────────────────────────────────────────────────
# 6. Gerenciar planilha Excel
# ─────────────────────────────────────────────────

def load_existing_data(filename: str) -> dict[str, list[dict]]:
    """Lê a planilha e devolve {cidade: [records]}."""
    if not os.path.exists(filename):
        return {}

    wb = openpyxl.load_workbook(filename, data_only=True)
    ws = wb.active
    data: dict[str, list[dict]] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        cidade = str(row[0]).strip()
        orgao  = str(row[1] or "").strip()
        email  = str(row[2] or "").strip()
        data.setdefault(cidade, []).append(
            {"cidade": cidade, "orgao": orgao, "email": email}
        )

    log.info("Planilha carregada: %d cidades com dados existentes.", len(data))
    return data


def are_records_equal(old: list[dict], new: list[dict]) -> bool:
    """Compara duas listas de registros (ordem-independente)."""
    if len(old) != len(new):
        return False

    def key(r):
        return (r.get("orgao", ""), r.get("email", ""))

    for o, n in zip(sorted(old, key=key), sorted(new, key=key)):
        if o["orgao"] != n["orgao"] or o["email"] != n["email"]:
            return False
    return True


def init_excel_file(filename: str) -> None:
    """Cria planilha com cabeçalho formatado."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TJSC Emails"

    header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    header_fill  = PatternFill("solid", fgColor="005B96")  # Azul TJSC
    header_align = Alignment(horizontal="center", vertical="center")
    headers = ["Cidade", "Órgão", "E-mail"]
    widths  = [40, 55, 45]

    for col_idx, (hdr, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:C1"
    wb.save(filename)
    log.info("Planilha '%s' criada.", filename)


def upsert_records_in_excel(
    all_records: list[dict],
    filename: str,
    existing_data: dict[str, list[dict]],
) -> int:
    """
    Insere ou atualiza registros na planilha.
    Agrupa por cidade, compara com dados existentes, e salva incrementalmente.
    Retorna o número de cidades atualizadas.
    """
    if not os.path.exists(filename):
        init_excel_file(filename)

    # Agrupa novos registros por cidade
    by_city: dict[str, list[dict]] = {}
    for rec in all_records:
        by_city.setdefault(rec["cidade"], []).append(rec)

    updates = 0
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    alt_fill   = PatternFill("solid", fgColor="E8F4FC")
    plain_fill = PatternFill("solid", fgColor="FFFFFF")

    for cidade_label, new_recs in sorted(by_city.items()):
        old_recs = existing_data.get(cidade_label, [])

        if old_recs and are_records_equal(old_recs, new_recs):
            log.info("  [%s] Sem alterações (%d registros).", cidade_label, len(new_recs))
            continue

        # Remove linhas antigas desta cidade
        for row_idx in range(ws.max_row, 1, -1):
            val = ws.cell(row=row_idx, column=1).value
            if val and str(val).strip() == cidade_label:
                ws.delete_rows(row_idx)

        # Insere novos registros
        start_row = ws.max_row + 1
        for i, rec in enumerate(new_recs):
            r = start_row + i
            fill = alt_fill if r % 2 == 0 else plain_fill
            for col_idx, fld in enumerate(["cidade", "orgao", "email"], start=1):
                cell = ws.cell(row=r, column=col_idx, value=rec[fld])
                cell.fill      = fill
                cell.alignment = Alignment(vertical="center")

        if old_recs:
            log.info("  [%s] Atualizado: %d → %d registros.", cidade_label, len(old_recs), len(new_recs))
        else:
            log.info("  [%s] Inserido: %d registros.", cidade_label, len(new_recs))
        updates += 1

    ws.auto_filter.ref = f"A1:C{max(2, ws.max_row)}"
    wb.save(filename)
    return updates


# ─────────────────────────────────────────────────
# 7. Main
# ─────────────────────────────────────────────────

def main():
    log.info("=" * 65)
    log.info("TJSC - Consulta de E-mails — Scraper")
    log.info("=" * 65)
    log.info("Termos a buscar: %s", ", ".join(SEARCH_TERMS))
    log.info("Saída: %s", os.path.abspath(OUTPUT_FILE))
    log.info("=" * 65)

    # Planilha existente ou nova
    if os.path.exists(OUTPUT_FILE):
        log.info("Planilha existente encontrada — modo atualização.")
        existing_data = load_existing_data(OUTPUT_FILE)
    else:
        init_excel_file(OUTPUT_FILE)
        existing_data = {}

    driver = create_driver()
    all_records: list[dict] = []

    try:
        # ── Passo 1: Obter token reCAPTCHA ─────────────────────────────────
        recaptcha_token = get_recaptcha_token(driver)

        if not recaptcha_token:
            log.warning(
                "Token não capturado automaticamente. "
                "Tentando usar o browser para buscar diretamente..."
            )

        # ── Passo 2: Buscar cada termo ──────────────────────────────────────
        session = requests.Session()
        session.headers.update({
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/121.0.0.0 Safari/537.36"
            )
        })

        for term in SEARCH_TERMS:
            log.info("Buscando: '%s'", term)
            raw_results = []

            if recaptcha_token:
                # Tenta via API direta (mais rápido e estável)
                raw_results = search_api(term, recaptcha_token, session)
                
                # Se a API retornar erro de autenticação, tenta via browser
                if not raw_results:
                    log.warning(
                        "  API sem resultados para '%s'. Tentando via browser...", term
                    )
                    raw_results = search_via_browser(driver, term)
            else:
                # Fallback: busca direto pelo browser (requer user interaction p/ cada busca)
                raw_results = search_via_browser(driver, term)

            # Processa resultados
            parsed = parse_results(raw_results)
            log.info("  → %d registros válidos após filtragem.", len(parsed))
            all_records.extend(parsed)

            if term != SEARCH_TERMS[-1]:
                time.sleep(DELAY_BETWEEN_TERMS)

        # ── Passo 3: Salvar na planilha ────────────────────────────────────
        if all_records:
            log.info("Total de registros coletados: %d", len(all_records))
            updates = upsert_records_in_excel(all_records, OUTPUT_FILE, existing_data)
            log.info("=" * 65)
            log.info("Concluído! %d cidade(s) inseridas/atualizadas.", updates)
        else:
            log.warning("Nenhum registro coletado. Verifique o CAPTCHA e tente novamente.")

        log.info("Planilha: %s", os.path.abspath(OUTPUT_FILE))
        log.info("=" * 65)

    finally:
        try:
            input("\nPressione ENTER para fechar o browser e encerrar o script...")
        except EOFError:
            pass
        driver.quit()


if __name__ == "__main__":
    main()
