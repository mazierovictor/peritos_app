"""
TJRS - Comarcas - Web Scraper
==============================
Extrai dados de todas as comarcas do Estado do Rio Grande do Sul a partir de:
https://www.tjrs.jus.br/institu/comarcas/

Gera uma planilha Excel com as colunas:
  Cidade | Órgão | Email

Regras:
  - Emails de cabeçalho (Distribuição, Contadoria) são salvos com seu rótulo
  - Se "Vara Judicial" e "Direção do Foro" tiverem o MESMO e-mail, apenas
    "Vara Judicial" é salvo (de-duplicação)
  - Porto Alegre (e outras cidades com múltiplos foros) percorre todos os foros.
    Cada foro é identificado como "Cidade (Nome do Foro)"
  - A cada comarca processada, os dados são salvos incrementalmente na planilha
  - Em reexecuções, o script compara os dados e só atualiza se houver diferença

Dependências:
    pip install selenium webdriver-manager openpyxl
"""

import os
import re
import time
import logging

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    StaleElementReferenceException,
)

try:
    from webdriver_manager.chrome import ChromeDriverManager
    USE_WDM = True
except ImportError:
    USE_WDM = False

# ─────────────────────────────────────────────────
# Configurações
# ─────────────────────────────────────────────────
BASE_URL     = "https://www.tjrs.jus.br/institu/comarcas/"
OUTPUT_FILE  = "tjrs_guia_judiciario.xlsx"
DELAY_BETWEEN_COMARCAS = 2   # segundos entre comarcas
DELAY_CLICK  = 1.5           # segundos após clicar em cada vara
HEADLESS     = True          # False para abrir o browser e depurar
PAGE_TIMEOUT = 30

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# Regex para extrair endereço de e-mail
EMAIL_RE = re.compile(r'[\w.\-+]+@[\w.\-]+\.\w{2,}')

# Regex para extrair o índice da chamada ObtemDadosVaras(id, index)
VARA_INDEX_RE = re.compile(r'ObtemDadosVaras\s*\(\s*\d+\s*,\s*(\d+)\s*\)')

# Órgãos permitidos (palavras-chave em minúsculo sem acento)
# Qualquer vara/órgão cujo nome NÃO contenha alguma dessas palavras será ignorado.
ALLOWED_ORGANS = [
    "vara judicial",
    "vara civel",
    "vara de fazenda",
    "fazenda publica",
    "direcao do foro",
    "direcao de foro",
    "protocolo",
    "distribuicao",
    "contadoria",
]


def _normalize(text: str) -> str:
    """Remove acentos e converte para minúsculas para facilitar a comparação."""
    import unicodedata
    text = unicodedata.normalize('NFD', text)
    text = text.encode('ascii', 'ignore').decode('ascii')
    return text.lower()


def is_organ_allowed(name: str) -> bool:
    """Retorna True se o nome do órgão corresponde a alguma palavra-chave permitida."""
    norm = _normalize(name)
    return any(kw in norm for kw in ALLOWED_ORGANS)


# ─────────────────────────────────────────────────
# 1. Inicializar o WebDriver
# ─────────────────────────────────────────────────

def create_driver() -> webdriver.Chrome:
    opts = Options()
    if HEADLESS:
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/121.0.0.0 Safari/537.36"
    )

    if USE_WDM:
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=opts)
    else:
        driver = webdriver.Chrome(options=opts)

    driver.implicitly_wait(3)
    return driver


# ─────────────────────────────────────────────────
# 2. Helpers de extração
# ─────────────────────────────────────────────────

def _extract_email(text: str) -> str:
    """Retorna o primeiro endereço de e-mail encontrado no texto, ou ''."""
    m = EMAIL_RE.search(text)
    return m.group(0).lower().strip() if m else ""


def _extract_header_emails(body_text: str) -> list[dict]:
    """
    Extrai emails de cabeçalho da página (Distribuição, Contadoria, etc.).
    Exemplos de linhas:
        'Email Distribuição e Contadoria: fragudodistcont@tjrs.jus.br'
        'Email Distribuição: frpoacentdist@tjrs.jus.br'
        'Email Contadoria: frpoacentcontciv@tjrs.jus.br'
    Retorna lista de dicts {"orgao": ..., "email": ...}
    """
    records = []
    # Padrão: linha que contém "email" + (rótulo variável) + ":" + endereço
    pattern = re.compile(
        r'(Email\s+[\w\s]+?)\s*:\s*([\w.\-+]+@[\w.\-]+\.\w{2,})',
        re.IGNORECASE
    )
    for match in pattern.finditer(body_text):
        label = match.group(1).strip()
        email = match.group(2).lower().strip()
        # Normaliza o rótulo para algo legível
        label_normalized = re.sub(r'\s+', ' ', label).title()
        records.append({"orgao": label_normalized, "email": email})
    return records


# ─────────────────────────────────────────────────
# 3. Extrair dados de um foro já carregado
# ─────────────────────────────────────────────────

def scrape_foro(driver: webdriver.Chrome, cidade_label: str) -> list[dict]:
    """
    Extrai todos os e-mails da página de resultados atualmente carregada.
    Retorna lista de dicts {"cidade", "orgao", "email"}.
    """
    records: list[dict] = []
    body_text = driver.find_element(By.TAG_NAME, "body").text

    # ── Emails de cabeçalho (Distribuição, Contadoria etc.) ────────────────
    header_emails = _extract_header_emails(body_text)
    for rec in header_emails:
        records.append({
            "cidade": cidade_label,
            "orgao":  rec["orgao"],
            "email":  rec["email"],
        })

    # ── Varas / Órgãos (links ObtemDadosVaras) ────────────────────────────
    vara_links_data: list[tuple[int, str, str]] = []  # (index, text, href)
    try:
        links = driver.find_elements(By.XPATH, "//a[contains(@href,'ObtemDadosVaras')]")
        for link in links:
            href = link.get_attribute("href") or ""
            name = link.text.strip()
            idx_m = VARA_INDEX_RE.search(href)
            if name and idx_m and is_organ_allowed(name):
                vara_links_data.append((int(idx_m.group(1)), name, href))
            elif name and idx_m:
                log.debug("    [filtrado] %s", name)
    except Exception as exc:
        log.debug("  Erro ao listar links de varas: %s", exc)

    vara_emails: dict[str, str] = {}  # orgao_name → email

    for var_index, var_name, var_href in vara_links_data:
        try:
            # Re-localiza e clica
            links = driver.find_elements(By.XPATH, "//a[contains(@href,'ObtemDadosVaras')]")
            clicked = False
            for lnk in links:
                if lnk.text.strip() == var_name:
                    driver.execute_script("arguments[0].scrollIntoView(true);", lnk)
                    driver.execute_script("arguments[0].click();", lnk)
                    clicked = True
                    break

            if not clicked:
                # Tenta executar a chamada JS diretamente
                js_call = re.sub(r'javascript:', '', var_href)
                driver.execute_script(js_call)

            time.sleep(DELAY_CLICK)

            # O painel usa o ÍNDICE passado na chamada JS
            detail_id = f"tblDadosVaras{var_index}"
            email = ""
            try:
                detail_div = driver.find_element(By.ID, detail_id)
                panel_text = detail_div.text
                email = _extract_email(panel_text)

                if not email:
                    # Tenta via mailto: no innerHTML
                    inner = detail_div.get_attribute("innerHTML") or ""
                    m = re.search(r'mailto:([\w.\-+]+@[\w.\-]+\.\w{2,})', inner, re.I)
                    if m:
                        email = m.group(1).lower().strip()
            except NoSuchElementException:
                # Fallback: pega qualquer painel que tenha e-mail
                try:
                    panels = driver.find_elements(
                        By.XPATH,
                        "//div[starts-with(@id,'tblDadosVaras') and contains(text(),'@')]"
                    )
                    for p in panels:
                        email = _extract_email(p.text)
                        if email:
                            break
                except Exception:
                    pass

            if email:
                vara_emails[var_name] = email
                log.debug("    %s → %s", var_name, email)

        except StaleElementReferenceException:
            log.debug("  Link stale para %s (idx=%d), pulando", var_name, var_index)
        except Exception as exc:
            log.debug("  Erro ao processar %s: %s", var_name, exc)

    # ── Regra de de-duplicação ─────────────────────────────────────────────
    vara_jud  = vara_emails.get("Vara Judicial", "")
    dir_foro  = vara_emails.get("Direção do Foro", "")
    if vara_jud and dir_foro and vara_jud == dir_foro:
        del vara_emails["Direção do Foro"]
        log.debug("  De-dup: Vara Judicial == Direção do Foro → mantendo só Vara Judicial")

    for orgao_name, email in vara_emails.items():
        records.append({
            "cidade": cidade_label,
            "orgao":  orgao_name,
            "email":  email,
        })

    return records


# ─────────────────────────────────────────────────
# 4. Processar uma comarca (simples ou com sub-foros)
# ─────────────────────────────────────────────────

def _wait_for_comarca_load(driver: webdriver.Chrome, timeout: int = PAGE_TIMEOUT) -> bool:
    """Aguarda o carregamento do resultado da comarca (sumiu o select ou apareceu conteúdo)."""
    try:
        WebDriverWait(driver, timeout).until(
            EC.any_of(
                EC.presence_of_element_located((By.XPATH, "//a[contains(@href,'ObtemDadosVaras')]")),
                EC.presence_of_element_located((By.ID, "ForosRegionais")),
            )
        )
        return True
    except TimeoutException:
        return False


def scrape_comarca(driver: webdriver.Chrome, comarca_value: str, comarca_text: str) -> list[dict]:
    """
    Seleciona a comarca no dropdown da página principal e extrai todos os dados.
    Para comarcas com ForosRegionais (Porto Alegre etc.), itera sobre todos.
    """
    # Garante que estamos na página base
    if BASE_URL not in driver.current_url:
        driver.get(BASE_URL)

    wait = WebDriverWait(driver, PAGE_TIMEOUT)
    wait.until(EC.presence_of_element_located((By.ID, "Comarca")))

    comarca_sel = Select(driver.find_element(By.ID, "Comarca"))
    comarca_sel.select_by_value(comarca_value)
    time.sleep(2)

    # Verifica se carregou algum resultado
    _wait_for_comarca_load(driver, timeout=10)

    all_records: list[dict] = []

    # ── Sub-foros (Porto Alegre e similares) ──────────────────────────────
    has_foros = False
    try:
        foros_elem = driver.find_element(By.ID, "ForosRegionais")
        has_foros = True
    except NoSuchElementException:
        has_foros = False

    if has_foros:
        foros_sel  = Select(foros_elem)
        foro_opts  = [
            (opt.get_attribute("value").strip(), opt.text.strip())
            for opt in foros_sel.options
            if opt.get_attribute("value").strip()
               and not opt.text.strip().lower().startswith("selecione")
        ]
        log.info("  → %d foros em %s", len(foro_opts), comarca_text)

        # Nome base da comarca (ex: "Comarca de Porto Alegre" → "Porto Alegre")
        base_name = comarca_text
        if base_name.lower().startswith("comarca de "):
            base_name = base_name[len("comarca de "):]

        for foro_value, foro_text in foro_opts:
            cidade_label = f"{base_name} ({foro_text})"
            log.info("    Foro: %s", foro_text)

            # Seleciona o foro
            try:
                foros_elem = wait.until(
                    EC.presence_of_element_located((By.ID, "ForosRegionais"))
                )
                foros_sel = Select(foros_elem)
                foros_sel.select_by_value(foro_value)
            except Exception:
                # Recomeça a seleção completa se o elemento sumiu
                driver.get(BASE_URL)
                wait.until(EC.presence_of_element_located((By.ID, "Comarca")))
                comarca_sel = Select(driver.find_element(By.ID, "Comarca"))
                comarca_sel.select_by_value(comarca_value)
                time.sleep(2)
                foros_elem = wait.until(
                    EC.presence_of_element_located((By.ID, "ForosRegionais"))
                )
                foros_sel = Select(foros_elem)
                foros_sel.select_by_value(foro_value)

            time.sleep(2)
            _wait_for_comarca_load(driver, timeout=10)

            recs = scrape_foro(driver, cidade_label)
            log.info("    → %d registros.", len(recs))
            all_records.extend(recs)
            time.sleep(DELAY_BETWEEN_COMARCAS)

    else:
        # Comarca simples
        cidade_label = comarca_text
        if cidade_label.lower().startswith("comarca de "):
            cidade_label = cidade_label[len("comarca de "):]

        recs = scrape_foro(driver, cidade_label)
        log.info("  → %d registros.", len(recs))
        all_records.extend(recs)

    return all_records


# ─────────────────────────────────────────────────
# 5. Gerenciar planilha Excel
# ─────────────────────────────────────────────────

def load_existing_data(filename: str) -> dict[str, list[dict]]:
    """Lê a planilha e devolve {cidade: [records]}."""
    if not os.path.exists(filename):
        return {}

    wb = openpyxl.load_workbook(filename, data_only=True)
    ws = wb.active
    data: dict[str, list[dict]] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        cidade = str(row[0]).strip()
        orgao  = str(row[1] or "").strip()
        email  = str(row[2] or "").strip()
        data.setdefault(cidade, []).append(
            {"cidade": cidade, "orgao": orgao, "email": email}
        )

    log.info("Planilha carregada: %d entradas existentes.", len(data))
    return data


def are_records_equal(old: list[dict], new: list[dict]) -> bool:
    """Compara duas listas (ordem-independente, ignora campo 'cidade')."""
    if len(old) != len(new):
        return False

    def key(r):
        return (r.get("orgao", ""), r.get("email", ""))

    for o, n in zip(sorted(old, key=key), sorted(new, key=key)):
        if o["orgao"] != n["orgao"] or o["email"] != n["email"]:
            return False
    return True


def init_excel_file(filename: str) -> None:
    """Cria planilha com cabeçalho formatado."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TJRS Comarcas"

    header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    header_fill  = PatternFill("solid", fgColor="003366")
    header_align = Alignment(horizontal="center", vertical="center")
    headers = ["Cidade", "Órgão", "E-mail"]
    widths  = [40, 45, 45]

    for col_idx, (hdr, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:C1"
    wb.save(filename)
    log.info("Planilha '%s' criada.", filename)


def update_city_in_excel(data: list[dict], filename: str, cidade_label: str) -> None:
    """Remove linhas antigas da cidade e insere os novos registros ao final."""
    if not os.path.exists(filename):
        init_excel_file(filename)

    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    # Remove linhas antigas (de baixo p/ cima)
    for row_idx in range(ws.max_row, 1, -1):
        val = ws.cell(row=row_idx, column=1).value
        if val and str(val).strip() == cidade_label:
            ws.delete_rows(row_idx)

    # Insere novos registros
    alt_fill   = PatternFill("solid", fgColor="E8F0FE")
    plain_fill = PatternFill("solid", fgColor="FFFFFF")
    start_row  = ws.max_row + 1

    for i, rec in enumerate(data):
        r = start_row + i
        fill = alt_fill if r % 2 == 0 else plain_fill
        for col_idx, key in enumerate(["cidade", "orgao", "email"], start=1):
            cell = ws.cell(row=r, column=col_idx, value=rec[key])
            cell.fill      = fill
            cell.alignment = Alignment(vertical="center")

    ws.auto_filter.ref = f"A1:C{max(2, ws.max_row)}"
    wb.save(filename)


# ─────────────────────────────────────────────────
# 6. Main
# ─────────────────────────────────────────────────

def main():
    log.info("=" * 60)
    log.info("TJRS Comarcas Scraper — iniciando")
    log.info("=" * 60)

    driver = create_driver()

    try:
        # Lista de comarcas
        log.info("Carregando lista de comarcas em %s …", BASE_URL)
        driver.get(BASE_URL)
        WebDriverWait(driver, PAGE_TIMEOUT).until(
            EC.presence_of_element_located((By.ID, "Comarca"))
        )

        sel = Select(driver.find_element(By.ID, "Comarca"))
        comarcas = [
            (opt.get_attribute("value").strip(), opt.text.strip())
            for opt in sel.options
            if opt.get_attribute("value").strip()
               and not opt.text.strip().lower().startswith("selecione")
        ]
        log.info("  → %d comarcas encontradas.", len(comarcas))

        if not comarcas:
            log.error("Nenhuma comarca. Encerrando.")
            return

        # Planilha existente ou nova
        if os.path.exists(OUTPUT_FILE):
            log.info("Planilha existente encontrada — modo atualização.")
            existing_data = load_existing_data(OUTPUT_FILE)
        else:
            init_excel_file(OUTPUT_FILE)
            existing_data = {}

        total = len(comarcas)
        updates = 0

        for i, (val, text) in enumerate(comarcas, start=1):
            log.info("[%d/%d] %s (id=%s)", i, total, text, val)

            try:
                records = scrape_comarca(driver, val, text)
            except Exception as exc:
                log.error("  Falha em %s: %s", text, exc)
                try:
                    driver.get(BASE_URL)
                    WebDriverWait(driver, PAGE_TIMEOUT).until(
                        EC.presence_of_element_located((By.ID, "Comarca"))
                    )
                except Exception:
                    pass
                time.sleep(DELAY_BETWEEN_COMARCAS)
                continue

            if not records:
                log.info("  → Sem registros.")
                time.sleep(DELAY_BETWEEN_COMARCAS)
                continue

            # Agrupa por cidade (comarcas simples têm sempre 1 cidade)
            by_city: dict[str, list[dict]] = {}
            for rec in records:
                by_city.setdefault(rec["cidade"], []).append(rec)

            for cidade_label, city_recs in by_city.items():
                old = existing_data.get(cidade_label, [])
                if old and are_records_equal(old, city_recs):
                    log.info("  [%s] Sem alterações.", cidade_label)
                else:
                    update_city_in_excel(city_recs, OUTPUT_FILE, cidade_label)
                    updates += 1
                    if old:
                        log.info(
                            "  [%s] Alterado: %d → %d registros.",
                            cidade_label, len(old), len(city_recs),
                        )
                    else:
                        log.info(
                            "  [%s] %d registros inseridos.",
                            cidade_label, len(city_recs),
                        )

            # Volta à página principal
            driver.get(BASE_URL)
            try:
                WebDriverWait(driver, PAGE_TIMEOUT).until(
                    EC.presence_of_element_located((By.ID, "Comarca"))
                )
            except TimeoutException:
                driver.get(BASE_URL)

            if i < total:
                time.sleep(DELAY_BETWEEN_COMARCAS)

        log.info("=" * 60)
        log.info("Concluído! %d comarca(s) atualizadas/inseridas.", updates)
        log.info("Planilha: %s", os.path.abspath(OUTPUT_FILE))
        log.info("=" * 60)

    finally:
        driver.quit()


if __name__ == "__main__":
    main()
