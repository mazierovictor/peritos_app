"""
TJMG - Guia Judiciário 1ª Instância - Web Scraper
===================================================
Extrai dados de todas as comarcas do Estado de Minas Gerais a partir de:
https://www8.tjmg.jus.br/servicos/gj/guia/primeira_instancia/pesquisa.do

Gera uma planilha Excel com as colunas:
  Cidade | Órgão | Email

Dependências:
    pip install requests beautifulsoup4 openpyxl
"""

import os
import time
import unicodedata
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import logging

# ──────────────────────────────────────────────
# Configurações
# ──────────────────────────────────────────────
SEARCH_URL  = "https://www8.tjmg.jus.br/servicos/gj/guia/primeira_instancia/pesquisa.do"
CONSULTA_URL = "https://www8.tjmg.jus.br/servicos/gj/guia/primeira_instancia/consulta.do"
OPC_CONSULTA = "6"          # 6 = "TODOS OS DADOS"
OUTPUT_FILE  = "tjmg_guia_judiciario.xlsx"
DELAY_BETWEEN_REQUESTS = 2  # segundos entre requisições (respeite o servidor)

# Palavras-chave permitidas (sem acento, minúsculas)
ALLOWED_ORGANS = [
    "secretaria",
    "forum",
    "administracao",
    "vara civel",
    "vara de fazenda publica",
    "vara de familia",
    "vara unica",
    "contadoria",
]

# ─── Override pela UI (não altera lógica; só substitui a lista se houver config salva) ───
try:
    import json as _json_ui
    with open("scraper_config.json", encoding="utf-8") as _f_ui:
        _UI_CFG = _json_ui.load(_f_ui)
    if isinstance(_UI_CFG.get("palavras_chave"), list) and _UI_CFG["palavras_chave"]:
        ALLOWED_ORGANS = [str(x) for x in _UI_CFG["palavras_chave"]]
except Exception:
    pass

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/121.0.0.0 Safari/537.36"
    ),
    "Referer": SEARCH_URL,
}

def normalize_text(text: str) -> str:
    """Remove acentos e converte para minúsculas para facilitar a busca."""
    text = unicodedata.normalize('NFD', text)
    text = text.encode('ascii', 'ignore').decode('utf-8')
    return text.lower()

def is_organ_allowed(orgao_name: str) -> bool:
    """Verifica se o nome do órgão contém alguma das palavras-chave permitidas."""
    norm_name = normalize_text(orgao_name)
    for keyword in ALLOWED_ORGANS:
        if keyword in norm_name:
            return True
    return False

# ──────────────────────────────────────────────
# 1. Buscar lista de cidades do formulário
# ──────────────────────────────────────────────

def get_city_options(session: requests.Session) -> list[tuple[str, str]]:
    """
    Carrega a página de pesquisa e extrai as opções do select `codigoComposto`.
    Retorna uma lista de tuplas (value, label), e.g. [("MG_5", "Açucena"), ...]
    """
    log.info("Carregando lista de municípios...")
    response = session.get(SEARCH_URL, headers=HEADERS, timeout=30)
    response.raise_for_status()

    soup = BeautifulSoup(response.text, "html.parser")
    select = soup.find("select", {"name": "codigoComposto"})
    if select is None:
        raise RuntimeError(
            "Não foi possível encontrar o campo 'codigoComposto' na página. "
            "Verifique se o site ainda está disponível."
        )

    options = []
    for opt in select.find_all("option"):
        value = opt.get("value", "").strip()
        label = opt.get_text(strip=True)
        # Ignora a opção placeholder (value="0" ou valor vazio ou texto de seleção)
        if value and value != "0" and not label.startswith("["):
            options.append((value, label))

    log.info("  → %d municípios encontrados.", len(options))
    return options


# ──────────────────────────────────────────────
# 2. Consultar dados de uma cidade
# ──────────────────────────────────────────────

def fetch_city_data(session: requests.Session, codigo_composto: str, city_label: str) -> list[dict]:
    """
    Envia a consulta para uma cidade e extrai os órgãos/emails.
    Realiza paginação automática. Filtra pelos órgãos desejados.
    Se não houver nenhum fórum, retorna uma linha de observação "Não há fórum".
    """
    rows = []
    seen_records = set()
    pagina_atual = 1
    
    # O nome oficial da cidade será extraído na primeira página
    city_name = city_label

    while True:
        params = {
            "codigoComposto": codigo_composto,
            "opcConsulta": OPC_CONSULTA,
            "paginaFlag": "forum",
            "paginaForum": "1",
            "paginaJuizado": "1",
            "pagina": str(pagina_atual),
        }

        max_retries = 3
        retry_delay = 10
        success = False
        
        for attempt in range(max_retries):
            try:
                resp = session.get(CONSULTA_URL, params=params, headers=HEADERS, timeout=30)
                resp.raise_for_status()
                success = True
                break
            except requests.RequestException as exc:
                if attempt < max_retries - 1:
                    log.warning("  Erro ao consultar %s (%s) p.%d: %s. Tentando novamente em %ds (Tentativa %d/%d)...",
                                city_label, codigo_composto, pagina_atual, exc, retry_delay, attempt + 2, max_retries)
                    time.sleep(retry_delay)
                else:
                    log.error("  Falha ao consultar %s (%s) p.%d após %d tentativas: %s",
                              city_label, codigo_composto, pagina_atual, max_retries, exc)
                    
        if not success:
            break

        soup = BeautifulSoup(resp.text, "html.parser")

        # Na primeira página, tentamos capturar o nome oficial da cidade do cabeçalho
        if pagina_atual == 1:
            bold_tag = soup.find("b")
            if bold_tag:
                raw = bold_tag.get_text(" ", strip=True)
                # Remove prefixo numérico (ex: "0005 Açucena" → "Açucena")
                parts = raw.split()
                if parts and parts[0].isdigit():
                    city_name = " ".join(parts[1:]).strip()
                elif raw:
                    city_name = raw

        # Linhas com classe "titulo_tabela" contêm os órgãos (3 colunas)
        page_rows_found = False
        new_records_on_page = 0
        
        for tr in soup.find_all("tr", class_="titulo_tabela"):
            # Apenas linhas com e-mail são as de órgãos judiciários válidas
            span_email = tr.find("span", class_="email")
            if span_email is None:
                continue

            page_rows_found = True

            tds = tr.find_all("td")
            orgao = tds[0].get_text(strip=True) if tds else ""
            email = span_email.get_text(strip=True)

            record_key = (orgao, email)
            if record_key not in seen_records:
                seen_records.add(record_key)
                new_records_on_page += 1

                if orgao and is_organ_allowed(orgao):
                    rows.append({
                        "cidade": city_name,
                        "orgao": orgao,
                        "email": email,
                    })

        # Se não encontramos nenhuma nova linha válida (todas eram repetidas)
        # ou se a página não tem nenhuma linha, chegamos ao fim.
        if not page_rows_found or new_records_on_page == 0:
            break

        # Continua para a próxima página
        pagina_atual += 1
        time.sleep(DELAY_BETWEEN_REQUESTS)

    # Se após processar todas as páginas não houver nenum registro em `rows`
    # (seja porque a cidade não tem fórum, ou porque nenhum órgão atendeu aos filtros),
    # adicionamos o registro "Não há fórum" conforme solicitado.
    if not rows:
        rows.append({
            "cidade": city_label,
            "orgao": "Não há fórum",
            "email": "",
        })

    return rows


# ──────────────────────────────────────────────
# 3. Gerar planilha Excel (Sincronização Contínua)
# ──────────────────────────────────────────────

def load_existing_data(filename: str) -> dict[str, list[dict]]:
    """Lê a planilha existente e retorna os dados agrupados por cidade para comparação."""
    if not os.path.exists(filename):
        return {}
    
    wb = openpyxl.load_workbook(filename, data_only=True)
    ws = wb.active
    data = {}
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        cidade = str(row[0]).strip()
        orgao = str(row[1] or "").strip()
        email = str(row[2] or "").strip()
        
        if cidade not in data:
            data[cidade] = []
        data[cidade].append({"cidade": cidade, "orgao": orgao, "email": email})
        
    log.info("Planilha carregada: dados antigos estruturados para %d municípios.", len(data))
    return data

def are_records_equal(old_records: list[dict], new_records: list[dict]) -> bool:
    """Compara as listas de órgãos raspadas com as existentes na planilha."""
    if len(old_records) != len(new_records):
        return False
        
    def sort_key(r): return (r.get("orgao", ""), r.get("email", ""))
    
    for o, n in zip(sorted(old_records, key=sort_key), sorted(new_records, key=sort_key)):
        if o["orgao"] != n["orgao"] or o["email"] != n["email"] or o["cidade"] != n["cidade"]:
            return False
            
    return True

def init_excel_file(filename: str) -> None:
    """Cria a planilha inicial com os cabeçalhos formatados."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Guia Judiciário TJMG"

    header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    header_fill  = PatternFill("solid", fgColor="003366")  # azul escuro
    header_align = Alignment(horizontal="center", vertical="center")

    headers = ["Cidade", "Órgão", "E-mail"]
    column_widths = [30, 65, 40]

    for col_idx, (header_text, width) in enumerate(zip(headers, column_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:C1"

    wb.save(filename)
    log.info("Planilha '%s' inicializada.", filename)

def update_city_in_excel(data: list[dict], filename: str, city_name_to_remove: str) -> None:
    """Atualiza a planilha removendo os dados antigos da cidade (se houverem) e adicionando os novos."""
    if not os.path.exists(filename):
        init_excel_file(filename)
        
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    # 1. Encontrar e remover linhas antigas da cidade (de baixo para cima p/ não alterar índices)
    rows_to_delete = []
    for row_idx in range(ws.max_row, 1, -1):
        cell_val = ws.cell(row=row_idx, column=1).value
        city_in_sheet = str(cell_val).strip() if cell_val else ""
        if city_in_sheet == city_name_to_remove:
            rows_to_delete.append(row_idx)
            
    for r in rows_to_delete:
        ws.delete_rows(r)
        
    # 2. Anexar novos dados ao final
    alt_fill   = PatternFill("solid", fgColor="E8F0FE")
    plain_fill = PatternFill("solid", fgColor="FFFFFF")

    start_row = ws.max_row + 1

    for i, record in enumerate(data):
        current_row = start_row + i
        # Formatação alternada para toda a planilha baseada na linha absoluta atual
        fill = alt_fill if current_row % 2 == 0 else plain_fill
        
        for col_idx, key in enumerate(["cidade", "orgao", "email"], start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=record[key])
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")

    # 3. Atualizar o auto-filtro e salvar
    ws.auto_filter.ref = f"A1:C{max(2, ws.max_row)}"
    wb.save(filename)


# ──────────────────────────────────────────────
# 4. Main
# ──────────────────────────────────────────────

def main():
    session = requests.Session()
    session.headers.update(HEADERS)

    # Passo 1 – lista de cidades
    cities = get_city_options(session)
    if not cities:
        log.error("Nenhum município encontrado. Encerrando.")
        return

    # Passo 2 – carregar planilha existente ou criar nova
    existing_data = {}
    if os.path.exists(OUTPUT_FILE):
        log.info("Verificando atualizações sobre a planilha existente ('%s')...", OUTPUT_FILE)
        existing_data = load_existing_data(OUTPUT_FILE)
    else:
        init_excel_file(OUTPUT_FILE)

    total = len(cities)
    records_saved = 0
    updates_made = 0

    # Passo 3 – consultar cada cidade e ATUALIZAR apenas se houver diferença
    for i, (codigo, label) in enumerate(cities, start=1):
        log.info("[%d/%d] %s (%s)", i, total, label, codigo)
        
        records = fetch_city_data(session, codigo, label)
        
        if records:
            # records[0]['cidade'] contém o nome oficial usado no Excel (ex: Açucena ou Abadia dos Dourados)
            city_name_in_records = records[0]["cidade"]
            old_records = existing_data.get(city_name_in_records, [])
            
            if old_records and are_records_equal(old_records, records):
                log.info("  → Sem alterações. Consulta idêntica à planilha.")
                records_saved += len(records)
            else:
                update_city_in_excel(records, OUTPUT_FILE, city_name_in_records)
                records_saved += len(records)
                updates_made += 1
                if old_records:
                    log.info("  → Dados alterados! Planilha atualizada. (%d registros em vez de %d)", len(records), len(old_records))
                else:
                    log.info("  → %d registros extraídos e inseridos novos na planilha.", len(records))

        if i < total:
            time.sleep(DELAY_BETWEEN_REQUESTS)

    log.info("Processo concluído! %d cidades sofreram alterações/inserção nesta execução.", updates_made)

if __name__ == "__main__":
    main()
