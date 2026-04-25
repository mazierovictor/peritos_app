"""
TJRO - Guia de Telefones e E-mails - Web Scraper
==================================================
Extrai dados de todos os municípios do Estado de Rondônia a partir de:
https://www.tjro.jus.br/rhtransparente/telefones

Gera uma planilha Excel com as colunas:
  Cidade | Órgão | Email

Regras:
  - O site possui proteção anti-bot (F5 BIG-IP), por isso usa Selenium
  - Seleciona cada município no dropdown e submete o formulário
  - Extrai setores e e-mails do HTML retornado
  - Salva incrementalmente na planilha a cada município
  - Em reexecuções, compara e só atualiza se houver diferença
  - Browser fica visível para o usuário resolver CAPTCHAs se necessário

Dependências:
    pip install selenium webdriver-manager openpyxl beautifulsoup4
"""

import os
import re
import time
import unicodedata
import logging

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from bs4 import BeautifulSoup

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException

try:
    from webdriver_manager.chrome import ChromeDriverManager
    USE_WDM = True
except ImportError:
    USE_WDM = False

# ──────────────────────────────────────────────
# Configurações
# ──────────────────────────────────────────────
BASE_URL     = "https://www.tjro.jus.br/rhtransparente/telefones"
OUTPUT_FILE  = "tjro_guia_judiciario.xlsx"
DELAY        = 2      # segundos entre municípios
HEADLESS     = False  # False = browser visível (para resolver CAPTCHAs)
PAGE_TIMEOUT = 30

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

EMAIL_RE = re.compile(r'[\w.\-+]+@[\w.\-]+\.\w{2,}')


# ──────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────

def clean_text(text: str) -> str:
    """Remove espaços extras e quebras de linha."""
    return " ".join(text.split())


def extract_email_from_text(text: str) -> str:
    """Retorna o primeiro e-mail encontrado, ou ''."""
    m = EMAIL_RE.search(text)
    return m.group(0).lower().strip() if m else ""


# ──────────────────────────────────────────────
# 1. Driver
# ──────────────────────────────────────────────

def create_driver() -> webdriver.Chrome:
    opts = Options()
    if HEADLESS:
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1400,900")
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/121.0.0.0 Safari/537.36"
    )
    if USE_WDM:
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=opts)
    else:
        driver = webdriver.Chrome(options=opts)
    driver.implicitly_wait(3)
    return driver


# ──────────────────────────────────────────────
# 2. Aguardar página / dropdown
# ──────────────────────────────────────────────

def wait_for_dropdown(driver: webdriver.Chrome, timeout: int = PAGE_TIMEOUT) -> bool:
    """Aguarda o dropdown de municípios. Se não aparecer, pode ser CAPTCHA."""
    try:
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.NAME, "comarcaLotacaoPredioId"))
        )
        return True
    except TimeoutException:
        return False


def load_main_page(driver: webdriver.Chrome) -> bool:
    """
    Carrega a página principal e aguarda o dropdown.
    Se cair em CAPTCHA, aguarda até 5 minutos para o usuário resolver.
    """
    driver.get(BASE_URL)
    if wait_for_dropdown(driver, timeout=20):
        return True

    # Pode ser CAPTCHA - aguarda o usuário resolver
    log.warning("Dropdown não encontrado. Possível CAPTCHA — aguardando até 5 min para resolução manual...")
    return wait_for_dropdown(driver, timeout=300)


# ──────────────────────────────────────────────
# 3. Lista de municípios
# ──────────────────────────────────────────────

def get_city_options(driver: webdriver.Chrome) -> list[tuple[str, str]]:
    """Retorna lista de (value, label) com todos os municípios do dropdown."""
    log.info("Carregando lista de municípios em %s …", BASE_URL)
    if not load_main_page(driver):
        log.error("Não foi possível carregar a lista de municípios.")
        return []

    sel_elem = driver.find_element(By.NAME, "comarcaLotacaoPredioId")
    sel = Select(sel_elem)
    options = []
    for opt in sel.options:
        value = opt.get_attribute("value").strip()
        text  = opt.text.strip()
        if value and not text.lower().startswith("selecione") and not text.startswith("-"):
            options.append((value, text))

    log.info("  → %d municípios encontrados.", len(options))
    return options


# ──────────────────────────────────────────────
# 4. Parsear HTML de resultados
# ──────────────────────────────────────────────

def parse_results_html(html: str, city_label: str) -> list[dict]:
    """
    Parseia o HTML da página de resultados do TJRO.

    Estrutura esperada (confirmada visualmente):
      <div class="style3 setorTitulo">
        <span> CODIGO - Nome do Órgão</span><br>
        <span class="style3 setorEmail">email@tjro.jus.br</span>
      </div>
      <div class="style3 telefone1">
        <span>(69) 3309-XXXX Descrição</span>
      </div>
      ...

    Também há entradas sem e-mail (apenas telefone).
    """
    soup = BeautifulSoup(html, "html.parser")
    records = []
    seen = set()

    # ── Nome real do município ────────────────────────────────────────────
    city_name = city_label
    titulo_tag = soup.find("h3", class_="titulo")
    if titulo_tag:
        raw = clean_text(titulo_tag.get_text())
        for prefix in ["Município de ", "Municipio de ", "Município: ", "Comarca de ", "Comarca: "]:
            if raw.startswith(prefix):
                raw = raw[len(prefix):].strip()
                break
        if raw:
            city_name = raw

    # ── Iterar sobre cada bloco setorTitulo ──────────────────────────────
    for setor_div in soup.find_all("div", class_="setorTitulo"):
        full_text = clean_text(setor_div.get_text())

        # Nome do órgão: primeiro span sem classe setorEmail
        orgao = ""
        email = ""

        for child in setor_div.children:
            # Pega apenas tags span
            if getattr(child, "name", None) != "span":
                continue
            classes = child.get("class") or []
            span_text = clean_text(child.get_text())
            if not span_text:
                continue

            if "setorEmail" in classes:
                # É o span do e-mail
                candidate = extract_email_from_text(span_text)
                if candidate:
                    email = candidate
                elif span_text == ".":
                    email = ""  # ponto significa sem e-mail
            else:
                # É o span do nome do órgão
                if not orgao:
                    orgao = span_text

        # Fallback: se não achou e-mail via span, tenta no HTML completo do div
        if not email:
            # Tenta <a href="mailto:...">
            mailto = setor_div.find("a", href=re.compile(r'^mailto:', re.I))
            if mailto:
                email = mailto.get("href", "")[7:].strip().lower()
            else:
                # Tenta qualquer e-mail no texto do div
                candidate = extract_email_from_text(full_text)
                if candidate:
                    email = candidate

        if not orgao:
            continue

        # Chave de deduplicação
        key = (orgao, email)
        if key in seen:
            continue
        seen.add(key)

        records.append({
            "cidade": city_name,
            "orgao":  orgao,
            "email":  email,
        })
        log.debug("    %s → %s", orgao, email)

    log.info("  → %d registros extraídos de %s.", len(records), city_label)

    if not records:
        # Verifica se há algum setorTitulo na página (mesmo sem e-mail)
        all_setores = soup.find_all("div", class_="setorTitulo")
        if not all_setores:
            records.append({
                "cidade": city_name,
                "orgao":  "Não há fórum",
                "email":  "",
            })

    return records


# ──────────────────────────────────────────────
# 5. Processar um município
# ──────────────────────────────────────────────

def scrape_city(driver: webdriver.Chrome, city_value: str, city_label: str) -> list[dict]:
    """
    Seleciona o município no dropdown, submete e extrai os dados.
    """
    # Garante que está na página principal
    if "pesquisa" in driver.current_url or "rhtransparente" not in driver.current_url:
        if not load_main_page(driver):
            return []
    else:
        # Verifica se o dropdown está presente
        try:
            driver.find_element(By.NAME, "comarcaLotacaoPredioId")
        except NoSuchElementException:
            if not load_main_page(driver):
                return []

    wait = WebDriverWait(driver, PAGE_TIMEOUT)

    # Seleciona o município
    try:
        sel_elem = wait.until(EC.presence_of_element_located((By.NAME, "comarcaLotacaoPredioId")))
        Select(sel_elem).select_by_value(city_value)
    except Exception as exc:
        log.error("  Erro ao selecionar %s: %s", city_label, exc)
        return []

    # Clica em Consultar (input type=image ou img com classe iconeConsultar)
    try:
        btn = driver.find_element(By.CSS_SELECTOR, "input[type='image'], img.iconeConsultar, img.imgButton")
        driver.execute_script("arguments[0].click();", btn)
    except NoSuchElementException:
        try:
            driver.find_element(By.TAG_NAME, "form").submit()
        except Exception as e:
            log.error("  Botão Consultar não encontrado para %s: %s", city_label, e)
            return []

    # Aguarda o carregamento dos resultados
    time.sleep(2)
    try:
        wait.until(EC.any_of(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div.setorTitulo")),
            EC.presence_of_element_located((By.CSS_SELECTOR, "h3.titulo")),
            EC.presence_of_element_located((By.CSS_SELECTOR, "h3.predioTitulo")),
        ))
    except TimeoutException:
        # Verifica se caiu em CAPTCHA novamente
        if "bloqueada" in driver.page_source.lower() or "captcha" in driver.page_source.lower():
            log.warning("  CAPTCHA detectado! Aguardando 5 min para resolução manual...")
            # Tenta voltar à página principal e re-processar após resolução
            time.sleep(300)
            return scrape_city(driver, city_value, city_label)
        log.warning("  Timeout aguardando resultados de %s. Tentando parsear…", city_label)

    html = driver.page_source
    records = parse_results_html(html, city_label)

    # Volta à página principal para a próxima iteração
    driver.get(BASE_URL)
    wait_for_dropdown(driver, timeout=20)

    return records


# ──────────────────────────────────────────────
# 6. Planilha Excel
# ──────────────────────────────────────────────

def load_existing_data(filename: str) -> dict[str, list[dict]]:
    if not os.path.exists(filename):
        return {}
    wb = openpyxl.load_workbook(filename, data_only=True)
    ws = wb.active
    data: dict[str, list[dict]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        cidade = str(row[0]).strip()
        orgao  = str(row[1] or "").strip()
        email  = str(row[2] or "").strip()
        data.setdefault(cidade, []).append({"cidade": cidade, "orgao": orgao, "email": email})
    log.info("Planilha carregada: %d cidades existentes.", len(data))
    return data


def are_records_equal(old: list[dict], new: list[dict]) -> bool:
    if len(old) != len(new):
        return False
    def key(r):
        return (r.get("orgao", ""), r.get("email", ""))
    for o, n in zip(sorted(old, key=key), sorted(new, key=key)):
        if o["orgao"] != n["orgao"] or o["email"] != n["email"]:
            return False
    return True


def init_excel_file(filename: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TJRO Guia Judiciário"
    header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    header_fill  = PatternFill("solid", fgColor="003366")
    header_align = Alignment(horizontal="center", vertical="center")
    headers = ["Cidade", "Órgão", "E-mail"]
    widths  = [30, 65, 45]
    for col_idx, (hdr, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:C1"
    wb.save(filename)
    log.info("Planilha '%s' criada.", filename)


def update_city_in_excel(data: list[dict], filename: str, city_name: str) -> None:
    if not os.path.exists(filename):
        init_excel_file(filename)
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    for row_idx in range(ws.max_row, 1, -1):
        val = ws.cell(row=row_idx, column=1).value
        if val and str(val).strip() == city_name:
            ws.delete_rows(row_idx)
    alt_fill   = PatternFill("solid", fgColor="E8F0FE")
    plain_fill = PatternFill("solid", fgColor="FFFFFF")
    start_row  = ws.max_row + 1
    for i, rec in enumerate(data):
        r = start_row + i
        fill = alt_fill if r % 2 == 0 else plain_fill
        for col_idx, key in enumerate(["cidade", "orgao", "email"], start=1):
            cell = ws.cell(row=r, column=col_idx, value=rec[key])
            cell.fill      = fill
            cell.alignment = Alignment(vertical="center")
    ws.auto_filter.ref = f"A1:C{max(2, ws.max_row)}"
    wb.save(filename)


# ──────────────────────────────────────────────
# 7. Main
# ──────────────────────────────────────────────

def main():
    log.info("=" * 60)
    log.info("TJRO Guia Judiciário Scraper — iniciando")
    log.info("Saída: %s", os.path.abspath(OUTPUT_FILE))
    log.info("=" * 60)

    driver = create_driver()

    try:
        # Lista de municípios
        cities = get_city_options(driver)
        if not cities:
            log.error("Nenhum município encontrado. Encerrando.")
            return

        # Planilha
        if os.path.exists(OUTPUT_FILE):
            log.info("Planilha existente — modo atualização.")
            existing_data = load_existing_data(OUTPUT_FILE)
        else:
            init_excel_file(OUTPUT_FILE)
            existing_data = {}

        total   = len(cities)
        updates = 0

        for i, (value, label) in enumerate(cities, start=1):
            log.info("[%d/%d] %s (id=%s)", i, total, label, value)

            try:
                records = scrape_city(driver, value, label)
            except Exception as exc:
                log.error("  Falha em %s: %s", label, exc)
                try:
                    driver.get(BASE_URL)
                    wait_for_dropdown(driver, timeout=20)
                except Exception:
                    pass
                time.sleep(DELAY)
                continue

            if not records:
                log.info("  → Sem registros.")
                time.sleep(DELAY)
                continue

            # Agrupa por cidade
            by_city: dict[str, list[dict]] = {}
            for rec in records:
                by_city.setdefault(rec["cidade"], []).append(rec)

            for city_name, city_recs in by_city.items():
                old = existing_data.get(city_name, [])
                if old and are_records_equal(old, city_recs):
                    log.info("  [%s] Sem alterações.", city_name)
                else:
                    update_city_in_excel(city_recs, OUTPUT_FILE, city_name)
                    updates += 1
                    if old:
                        log.info("  [%s] Alterado: %d → %d registros.", city_name, len(old), len(city_recs))
                    else:
                        log.info("  [%s] %d registros inseridos.", city_name, len(city_recs))

            if i < total:
                time.sleep(DELAY)

        log.info("=" * 60)
        log.info("Concluído! %d município(s) atualizado(s)/inserido(s).", updates)
        log.info("Planilha: %s", os.path.abspath(OUTPUT_FILE))
        log.info("=" * 60)

    finally:
        driver.quit()


if __name__ == "__main__":
    main()
