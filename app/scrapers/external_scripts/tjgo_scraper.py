"""
TJGO - Agenda Eletrônica (SIGO) - Web Scraper
==============================================
Coleta e-mails de órgãos jurisdicionais do TJGO a partir da API JSON pública:
  https://sigo-backend.tjgo.jus.br/api/agenda/publico/localidades

Gera tjgo_guia_judiciario.xlsx com as colunas: Cidade | Órgão | E-mail

Dependências:
    pip install requests openpyxl
"""
from __future__ import annotations

import time
import unicodedata
import logging

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ──────────────────────────────────────────────
# Configurações
# ──────────────────────────────────────────────
API_URL = "https://sigo-backend.tjgo.jus.br/api/agenda/publico/localidades"
OUTPUT_FILE = "tjgo_guia_judiciario.xlsx"
PAGE_SIZE = 1000              # o servidor limita ~2000 por página
DELAY_BETWEEN_REQUESTS = 1    # segundos entre páginas (respeite o servidor)
MAX_RETRIES = 4
RETRY_DELAY = 5              # segundos entre tentativas

# Allowlist jurisdicional (sem acento, minúsculas). Calibrada contra os dados
# reais: 'secretaria' fica DE FORA (captura secretarias administrativas);
# 'forum' fica DENTRO (senão perde os fóruns das comarcas).
ALLOWED_ORGANS = [
    "vara",
    "juizado",
    "jurisdicional",
    "cejusc",
    "turma recursal",
    "forum",
    "contadoria",
    "tribunal do juri",
    "auditoria militar",
]

# ─── Override pela UI (não altera a lógica; só substitui a lista se houver config) ───
try:
    import json as _json_ui
    with open("scraper_config.json", encoding="utf-8") as _f_ui:
        _UI_CFG = _json_ui.load(_f_ui)
    if isinstance(_UI_CFG.get("palavras_chave"), list) and _UI_CFG["palavras_chave"]:
        ALLOWED_ORGANS = [str(x) for x in _UI_CFG["palavras_chave"]]
except Exception:
    pass

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/121.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json",
}

# ──────────────────────────────────────────────
# Filtragem (idêntica aos demais scrapers)
# ──────────────────────────────────────────────
_EXC_CRIMINAL = (
    "criminal", "criminais", "crime", "penal", "penais", "penas",
    "execucao penal", "execucoes penais", "socioeducativ", "socieducativ",
    "do juri", "de juri", "juiz de garantias", "juizo de garantias",
)
_EXC_INFANCIA = ("infancia", "juventude")
_CIVEL_OVERRIDE = (
    "civel", "civil", "fazenda", "fiscal", "fiscais", "familia", "sucessoes", "orfaos",
    "empresarial", "unica", "unico", "jurisdicional", "precatoria", "divida",
    "falencia", "recupera", "acidente",
)


def normalize_text(text: str) -> str:
    """Remove acentos e converte para minúsculas para facilitar a busca."""
    text = unicodedata.normalize("NFD", text or "")
    text = text.encode("ascii", "ignore").decode("utf-8")
    return text.lower()


def _excluir_orgao(nome_norm: str) -> bool:
    """True se a unidade for exclusivamente criminal/penal ou de infância/juventude.
    Recebe o nome JÁ normalizado (minúsculo, sem acento)."""
    suspeito = (any(t in nome_norm for t in _EXC_CRIMINAL)
                or any(t in nome_norm for t in _EXC_INFANCIA))
    if not suspeito:
        return False
    return not any(t in nome_norm for t in _CIVEL_OVERRIDE)


def is_organ_allowed(orgao_name: str) -> bool:
    """Mantém qualquer vara/juizado/unidade jurisdicional (genérica inclusive) e
    os órgãos de apoio configurados em ALLOWED_ORGANS; descarta criminal/infância
    puros."""
    norm_name = normalize_text(orgao_name).strip()
    if _excluir_orgao(norm_name):
        return False
    if "vara" in norm_name or "juizado" in norm_name or "jurisdicional" in norm_name:
        return True
    return any(keyword in norm_name for keyword in ALLOWED_ORGANS)


# ──────────────────────────────────────────────
# Extração das linhas a partir das lotações da API
# ──────────────────────────────────────────────
def extract_rows(localidades: list[dict]) -> list[dict]:
    """Filtra as lotações pela política e mapeia para {cidade, orgao, email}.
    Quando a API retorna dois e-mails separados por '/' cria uma linha por e-mail.
    Todos os e-mails são normalizados para minúsculas."""
    rows: list[dict] = []
    for loc in localidades:
        raw_email = (loc.get("email") or "").strip()
        if not raw_email:
            continue
        nome = (loc.get("nome") or "").strip()
        if not nome or not is_organ_allowed(nome):
            continue
        predio = loc.get("predio") or {}
        cidade = (predio.get("cidade") or "").strip()
        emails = [e.strip().lower() for e in raw_email.split("/") if e.strip() and "@" in e.strip()]
        for email in emails:
            rows.append({"cidade": cidade, "orgao": nome, "email": email})
    return rows


# ──────────────────────────────────────────────
# Fetch paginado da API pública (com retry/backoff)
# ──────────────────────────────────────────────
def fetch_all_localidades(session: requests.Session) -> list[dict]:
    """Pagina por /agenda/publico/localidades até hasNext=False. Retorna a lista
    bruta de lotações. Levanta RuntimeError se uma página falhar após MAX_RETRIES."""
    todos: list[dict] = []
    page = 0
    while True:
        params = {"page": page, "size": PAGE_SIZE}
        payload = None
        for attempt in range(MAX_RETRIES):
            try:
                resp = session.get(API_URL, params=params, headers=HEADERS, timeout=60)
                resp.raise_for_status()
                payload = resp.json()
                break
            except (requests.RequestException, ValueError) as exc:
                if attempt < MAX_RETRIES - 1:
                    log.warning("Falha na página %d: %s. Retentando em %ds (%d/%d)...",
                                page, exc, RETRY_DELAY, attempt + 2, MAX_RETRIES)
                    time.sleep(RETRY_DELAY)
                else:
                    raise RuntimeError(
                        f"Falha ao buscar a página {page} após {MAX_RETRIES} tentativas: {exc}"
                    ) from exc

        data = payload.get("data") or []
        todos.extend(data)
        page_info = payload.get("page") or {}
        log.info("Página %d: %d registros (acumulado: %d/%s)",
                 page, len(data), len(todos), page_info.get("totalElements", "?"))

        if not page_info.get("hasNext"):
            break
        page += 1
        time.sleep(DELAY_BETWEEN_REQUESTS)

    return todos


# ──────────────────────────────────────────────
# Geração da planilha Excel
# ──────────────────────────────────────────────
def write_excel(rows: list[dict], filename: str) -> None:
    """Gera a planilha do zero com as colunas Cidade | Órgão | E-mail."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Guia Judiciário TJGO"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor="003366")
    header_align = Alignment(horizontal="center", vertical="center")

    headers = ["Cidade", "Órgão", "E-mail"]
    column_widths = [30, 65, 40]
    for col_idx, (header_text, width) in enumerate(zip(headers, column_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    alt_fill = PatternFill("solid", fgColor="E8F0FE")
    plain_fill = PatternFill("solid", fgColor="FFFFFF")

    for i, record in enumerate(rows, start=2):
        fill = alt_fill if i % 2 == 0 else plain_fill
        for col_idx, key in enumerate(["cidade", "orgao", "email"], start=1):
            cell = ws.cell(row=i, column=col_idx, value=record[key])
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")

    ws.auto_filter.ref = f"A1:C{max(2, ws.max_row)}"
    wb.save(filename)


# ──────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────
def main() -> None:
    session = requests.Session()
    session.headers.update(HEADERS)

    log.info("Buscando localidades públicas do TJGO...")
    localidades = fetch_all_localidades(session)
    log.info("Total bruto: %d lotações.", len(localidades))

    rows = extract_rows(localidades)
    log.info("Após filtragem: %d órgãos com e-mail.", len(rows))

    write_excel(rows, OUTPUT_FILE)
    log.info("Planilha '%s' gerada com %d linhas.", OUTPUT_FILE, len(rows))


if __name__ == "__main__":
    main()
