"""
TJRJ - Serventias 1ª Instância - Web Scraper
==============================================
Extrai dados de contato das serventias do Tribunal de Justiça
do Rio de Janeiro:
  https://www3.tjrj.jus.br/consultasportalWeb/#/consultas/endereco_telefones/serventias-1inst

Itera sobre todas as comarcas × atribuições selecionadas e
coleta: Cidade, Órgão, Email, Telefone.

O site NÃO exige CAPTCHA real — usa requests simples (sem Selenium).

Dependências:
    pip install requests openpyxl
"""

import os
import re
import time
import logging

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ─────────────────────────────────────────────────
# Configurações
# ─────────────────────────────────────────────────
BASE_API = "https://www3.tjrj.jus.br/consultasportal/api/v1/telefonesEnderecos/serventias1Inst"
OUTPUT_FILE = "tjrj_guia_judiciario.xlsx"

# Atribuições a buscar (label → código)
ATRIBUICOES = [
    ("CIVEL",                             18),
    ("CIVEL ESPECIALIZADA (PESSOAS IDOSAS)", 7620),
    ("DIVIDA ATIVA",                      41),
    ("EMPRESARIAL",                       14),
    ("FAZENDA PUBLICA",                   36),
    ("FORUM",                             51),
    ("JUIZADO ESPECIAL CIVEL",            56),
    ("JUIZADO ESPECIAL DE FAZENDA PUBLICA", 64),
    ("NUCLEO DE JUSTICA 4.0",             101),
]

DELAY_BETWEEN_REQUESTS = 2  # segundos entre cada requisição à API

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# Regex para extrair e-mail
EMAIL_RE = re.compile(r'[\w.\-+]+@[\w.\-]+\.\w{2,}', re.IGNORECASE)


# ─────────────────────────────────────────────────
# 1. API helpers
# ─────────────────────────────────────────────────

def _api_get(session: requests.Session, endpoint: str, params: dict | None = None) -> list | dict:
    """Faz GET na API do TJRJ e retorna o JSON."""
    url = f"{BASE_API}/{endpoint}"
    p = {"recaptcha": "skipRecap2020"}
    if params:
        p.update(params)
    try:
        resp = session.get(url, params=p, timeout=30)
        resp.raise_for_status()
        return resp.json()
    except Exception as exc:
        log.error("Erro ao acessar %s: %s", url, exc)
        return []


def fetch_comarcas(session: requests.Session) -> list[dict]:
    """Retorna lista de comarcas [{label, value}, ...]."""
    data = _api_get(session, "cboComarca")
    log.info("Comarcas encontradas: %d", len(data))
    return data


def fetch_serventias(
    session: requests.Session,
    codigo_orgao: str,
    codigo_regional: str,
    codigo_atribuicao: int,
) -> list[dict]:
    """Consulta serventias para uma comarca + atribuição."""
    params = {
        "codigoOrgao": codigo_orgao,
        "codigoRegional": codigo_regional,
        "codigoTipoServentia": "0",   # TODAS
        "codigoAtribuicao": str(codigo_atribuicao),
    }
    return _api_get(session, "listar", params)


# ─────────────────────────────────────────────────
# 2. Processar resultados
# ─────────────────────────────────────────────────

def parse_serventias(raw_results: list[dict]) -> list[dict]:
    """
    Converte os resultados brutos da API em registros:
      {cidade, orgao, email, telefone}
    """
    records = []
    for item in raw_results:
        if not isinstance(item, dict):
            continue

        cidade = str(item.get("cidade", "") or "").strip()
        orgao  = str(item.get("nomeVara", "") or item.get("serventia", "") or "").strip()
        email  = str(item.get("email", "") or "").strip().lower()
        telefone = str(item.get("secretaria", "") or "").strip()

        # Tenta campo 'outros' se secretaria vazio
        if not telefone:
            telefone = str(item.get("outros", "") or "").strip()

        # Comarca como fallback para cidade
        if not cidade:
            comarca = str(item.get("comarcaEscolhida", "") or "").strip()
            if comarca.startswith("COMARCA DE "):
                cidade = comarca[len("COMARCA DE "):]
            elif comarca.startswith("COMARCA DA "):
                cidade = comarca[len("COMARCA DA "):]
            elif comarca:
                cidade = comarca

        # Normaliza cidade para title case
        if cidade:
            cidade = cidade.strip().title()

        if not orgao:
            continue

        orgao = orgao.strip().title()

        # Sem email → registra mesmo assim com "Não informado"
        if not email or "@" not in email:
            email = "Não informado"

        records.append({
            "cidade":   cidade or "Não informado",
            "orgao":    orgao,
            "email":    email,
            "telefone": telefone or "Não informado",
        })

    return records


# ─────────────────────────────────────────────────
# 3. Gerenciar planilha Excel
# ─────────────────────────────────────────────────

HEADERS = ["Cidade", "Órgão", "E-mail", "Telefone"]
WIDTHS  = [40, 55, 45, 25]


def load_existing_data(filename: str) -> dict[str, list[dict]]:
    """Lê a planilha e devolve {cidade: [records]}."""
    if not os.path.exists(filename):
        return {}

    wb = openpyxl.load_workbook(filename, data_only=True)
    ws = wb.active
    data: dict[str, list[dict]] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        cidade   = str(row[0]).strip()
        orgao    = str(row[1] or "").strip()
        email    = str(row[2] or "").strip()
        telefone = str(row[3] or "").strip() if len(row) > 3 else ""
        data.setdefault(cidade, []).append(
            {"cidade": cidade, "orgao": orgao, "email": email, "telefone": telefone}
        )

    log.info("Planilha carregada: %d cidades com dados existentes.", len(data))
    return data


def are_records_equal(old: list[dict], new: list[dict]) -> bool:
    """Compara duas listas de registros (ordem-independente)."""
    if len(old) != len(new):
        return False

    def key(r):
        return (r.get("orgao", ""), r.get("email", ""), r.get("telefone", ""))

    for o, n in zip(sorted(old, key=key), sorted(new, key=key)):
        if key(o) != key(n):
            return False
    return True


def init_excel_file(filename: str) -> None:
    """Cria planilha com cabeçalho formatado."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TJRJ Serventias"

    header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    header_fill  = PatternFill("solid", fgColor="00457C")  # Azul escuro TJRJ
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, (hdr, w) in enumerate(zip(HEADERS, WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D1"
    wb.save(filename)
    log.info("Planilha '%s' criada.", filename)


def upsert_records_in_excel(
    all_records: list[dict],
    filename: str,
    existing_data: dict[str, list[dict]],
) -> int:
    """
    Insere ou atualiza registros na planilha.
    Agrupa por cidade, compara com dados existentes, e salva incrementalmente.
    Retorna o número de cidades atualizadas.
    """
    if not os.path.exists(filename):
        init_excel_file(filename)

    # Agrupa novos registros por cidade
    by_city: dict[str, list[dict]] = {}
    for rec in all_records:
        by_city.setdefault(rec["cidade"], []).append(rec)

    updates = 0
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    alt_fill   = PatternFill("solid", fgColor="E8F0F8")
    plain_fill = PatternFill("solid", fgColor="FFFFFF")

    for cidade_label, new_recs in sorted(by_city.items()):
        old_recs = existing_data.get(cidade_label, [])

        if old_recs and are_records_equal(old_recs, new_recs):
            log.info("  [%s] Sem alterações (%d registros).", cidade_label, len(new_recs))
            continue

        # Remove linhas antigas desta cidade
        for row_idx in range(ws.max_row, 1, -1):
            val = ws.cell(row=row_idx, column=1).value
            if val and str(val).strip() == cidade_label:
                ws.delete_rows(row_idx)

        # Insere novos registros
        start_row = ws.max_row + 1
        for i, rec in enumerate(new_recs):
            r = start_row + i
            fill = alt_fill if r % 2 == 0 else plain_fill
            for col_idx, fld in enumerate(["cidade", "orgao", "email", "telefone"], start=1):
                cell = ws.cell(row=r, column=col_idx, value=rec[fld])
                cell.fill      = fill
                cell.alignment = Alignment(vertical="center")

        if old_recs:
            log.info("  [%s] Atualizado: %d → %d registros.", cidade_label, len(old_recs), len(new_recs))
        else:
            log.info("  [%s] Inserido: %d registros.", cidade_label, len(new_recs))
        updates += 1

    ws.auto_filter.ref = f"A1:D{max(2, ws.max_row)}"
    wb.save(filename)
    return updates


# ─────────────────────────────────────────────────
# 4. Main
# ─────────────────────────────────────────────────

def main():
    log.info("=" * 65)
    log.info("TJRJ - Serventias 1ª Instância — Scraper")
    log.info("=" * 65)
    log.info("Atribuições: %d", len(ATRIBUICOES))
    log.info("Saída: %s", os.path.abspath(OUTPUT_FILE))
    log.info("=" * 65)

    # Planilha existente ou nova
    if os.path.exists(OUTPUT_FILE):
        log.info("Planilha existente encontrada — modo atualização.")
        existing_data = load_existing_data(OUTPUT_FILE)
    else:
        init_excel_file(OUTPUT_FILE)
        existing_data = {}

    # Sessão HTTP
    session = requests.Session()
    session.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/121.0.0.0 Safari/537.36"
        ),
        "Accept": "application/json, text/plain, */*",
        "Referer": "https://www3.tjrj.jus.br/consultasportalWeb/",
    })

    # Buscar comarcas
    comarcas = fetch_comarcas(session)
    if not comarcas:
        log.error("Não foi possível obter a lista de comarcas. Abortando.")
        return

    all_records: list[dict] = []
    seen_global: set[tuple] = set()  # de-duplicação global

    total_comarcas = len(comarcas)

    for idx_c, comarca in enumerate(comarcas, start=1):
        comarca_label = comarca["label"]
        codigo_orgao  = comarca["value"]

        # Para CAPITAL (406), Regional "TODAS" = 0
        # Para demais comarcas, Regional = codigoOrgao
        if str(codigo_orgao) == "406":
            codigo_regional = "0"
        else:
            codigo_regional = codigo_orgao

        log.info("─" * 65)
        log.info("[%d/%d] Comarca: %s (código: %s, regional: %s)",
                 idx_c, total_comarcas, comarca_label, codigo_orgao, codigo_regional)

        comarca_records = 0

        for atrib_label, atrib_code in ATRIBUICOES:
            time.sleep(DELAY_BETWEEN_REQUESTS)

            raw = fetch_serventias(session, codigo_orgao, codigo_regional, atrib_code)
            if not raw:
                log.info("  [%s] Nenhum resultado.", atrib_label)
                continue

            parsed = parse_serventias(raw)
            log.info("  [%s] %d serventias encontradas.", atrib_label, len(parsed))

            # De-duplicação global
            for rec in parsed:
                key = (rec["cidade"].lower(), rec["orgao"].lower(), rec["email"])
                if key not in seen_global:
                    seen_global.add(key)
                    all_records.append(rec)
                    comarca_records += 1

        log.info("  → %d registros novos para %s.", comarca_records, comarca_label)

        # Salvar incrementalmente a cada comarca
        if all_records:
            log.info("  [Salvando incrementalmente... %d registros até agora]", len(all_records))
            upsert_records_in_excel(all_records, OUTPUT_FILE, existing_data)

    # Salvar final
    if all_records:
        log.info("=" * 65)
        log.info("Total de registros coletados: %d", len(all_records))
        updates = upsert_records_in_excel(all_records, OUTPUT_FILE, existing_data)
        log.info("Concluído! %d cidade(s) inseridas/atualizadas.", updates)
    else:
        log.warning("Nenhum registro coletado. Verifique a conectividade.")

    log.info("Planilha: %s", os.path.abspath(OUTPUT_FILE))
    log.info("=" * 65)


if __name__ == "__main__":
    main()
